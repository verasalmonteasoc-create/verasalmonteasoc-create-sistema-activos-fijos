#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Importar catálogo con estructura correcta"""
import sys
sys.path.insert(0, '.')

from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
import os

# Conexión a BD
db_host = os.getenv('DB_HOST', 'localhost')
db_port = os.getenv('DB_PORT', '5434')
db_name = os.getenv('DB_NAME', 'asset_management')
db_user = os.getenv('DB_USER', 'postgres')
db_password = os.getenv('DB_PASSWORD', 'postgres123')

connection_string = f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
engine = create_engine(connection_string)
Session = sessionmaker(bind=engine)
session = Session()

from backend.models import ChartOfAccounts, AssetCategory

print("=" * 80)
print("IMPORTAR CATÁLOGO DE CUENTAS (ESTRUCTURA ESPECIAL)")
print("=" * 80)

# Cargar archivo
wb = load_workbook('Catalogo_Cuentas.xlsx')
ws = wb.active

print("\n[OK] Archivo cargado: Catalogo_Cuentas.xlsx")

# Procesar datos
accounts_dict = {}
row_num = 2

for row in ws.iter_rows(min_row=2, values_only=False):
    # Estructura: ActivoCode | ActivoDesc | GastoCode | GastoDesc | DeprecCode | DeprecDesc
    activo_code = row[0].value
    activo_desc = row[1].value
    gasto_code = row[2].value
    gasto_desc = row[3].value
    deprec_code = row[4].value
    deprec_desc = row[5].value

    if not activo_code:
        break

    # Agregar activo
    if activo_code not in accounts_dict:
        accounts_dict[str(activo_code).strip()] = {
            'name': str(activo_desc).strip() if activo_desc else '',
            'type': 'Activo',
            'description': ''
        }

    # Agregar gasto
    if gasto_code and gasto_code not in accounts_dict:
        accounts_dict[str(gasto_code).strip()] = {
            'name': str(gasto_desc).strip() if gasto_desc else 'Gasto Depreciación',
            'type': 'Gasto',
            'description': ''
        }

    # Agregar depreciación acumulada
    if deprec_code and deprec_code not in accounts_dict:
        accounts_dict[str(deprec_code).strip()] = {
            'name': str(deprec_desc).strip() if deprec_desc else 'Deprec. Acumulada',
            'type': 'Activo',  # Depreciación acumulada es contra-activo
            'description': ''
        }

    row_num += 1

print(f"\n[OK] {len(accounts_dict)} cuentas extraidas")
print("\nPrimeras 5 cuentas:")
for i, (code, data) in enumerate(list(accounts_dict.items())[:5], 1):
    print(f"  {i}. [{code}] {data['name']} ({data['type']})")

# Eliminar anterior
print(f"\n[INFO] Eliminando catálogo anterior...")
old_count = session.query(ChartOfAccounts).count()
session.query(ChartOfAccounts).delete()
session.commit()
print(f"  [OK] {old_count} cuentas eliminadas")

# Importar nuevas
print(f"\n[INFO] Importando {len(accounts_dict)} cuentas nuevas...")
for idx, (code, data) in enumerate(accounts_dict.items(), 1):
    account = ChartOfAccounts(
        code=code,
        name=data['name'],
        account_type=data['type'],
        description=data['description']
    )
    session.add(account)
    if idx % 50 == 0:
        print(f"  {idx} cuentas...")

session.commit()
print(f"  [OK] {len(accounts_dict)} cuentas importadas")

# Vincular a categorías
print(f"\n[INFO] Vinculando a categorías...")
categories = session.query(AssetCategory).all()

for category in categories:
    # Buscar gasto
    expense_account = session.query(ChartOfAccounts).filter(
        ChartOfAccounts.name.ilike('%Gasto%Deprec%')
    ).first()

    # Buscar depreciación acumulada
    acum_account = session.query(ChartOfAccounts).filter(
        ChartOfAccounts.name.ilike('%Deprec%Acumulada%')
    ).first()

    if expense_account:
        category.depreciation_expense_account = expense_account.code

    if acum_account:
        category.accumulated_depreciation_account = acum_account.code

session.commit()
print(f"  [OK] {len(categories)} categorías vinculadas")

print("\n" + "=" * 80)
print("[EXITO] IMPORTACION COMPLETADA")
print("=" * 80)
print(f"\nResumen:")
print(f"  Cuentas eliminadas: {old_count}")
print(f"  Cuentas importadas: {len(accounts_dict)}")
print(f"  Categorías vinculadas: {len(categories)}")

session.close()
