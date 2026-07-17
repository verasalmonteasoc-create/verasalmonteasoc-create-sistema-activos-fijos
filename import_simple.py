#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Script simple para importar catálogo de cuentas"""

import sys
import os

# Agregar el directorio actual al path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from openpyxl import load_workbook
    from dotenv import load_dotenv
    load_dotenv()

    # Conexión direc a la base de datos sin Flask
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    import os

    # Configurar conexión a DB
    db_host = os.getenv('DB_HOST', 'localhost')
    db_port = os.getenv('DB_PORT', '5434')
    db_name = os.getenv('DB_NAME', 'asset_management')
    db_user = os.getenv('DB_USER', 'postgres')
    db_password = os.getenv('DB_PASSWORD', 'postgres123')

    connection_string = f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}"
    engine = create_engine(connection_string)
    Session = sessionmaker(bind=engine)
    session = Session()

    # Importar modelos
    from backend.models import ChartOfAccounts, AssetCategory

    print("=" * 80)
    print("IMPORTAR CATÁLOGO DE CUENTAS")
    print("=" * 80)

    # Cargar archivo Excel
    filepath = 'Catalogo_Cuentas.xlsx'
    if not os.path.exists(filepath):
        print(f"✗ Archivo no encontrado: {filepath}")
        sys.exit(1)

    wb = load_workbook(filepath)
    ws = wb.active

    print(f"\n✓ Archivo cargado: {filepath}")
    print(f"  Sheet: {ws.title}")
    print(f"  Filas: {ws.max_row}")

    # Recolectar datos
    accounts_data = []
    errors = []
    row_num = 2

    print(f"\nLeyendo datos...")

    for row in ws.iter_rows(min_row=2, values_only=False):
        try:
            code = row[0].value if row[0] else None
            name = row[1].value if row[1] else None
            account_type = row[2].value if row[2] else None
            description = row[3].value if row[3] else None

            if not code or not name or not account_type:
                errors.append(f'Fila {row_num}: Faltan campos obligatorios')
                row_num += 1
                continue

            code = str(code).strip()
            name = str(name).strip()
            account_type = str(account_type).strip()
            description = str(description).strip() if description else ''

            valid_types = ['Activo', 'Pasivo', 'Capital', 'Ingreso', 'Gasto']
            if account_type not in valid_types:
                errors.append(f'Fila {row_num}: Tipo "{account_type}" no válido')
                row_num += 1
                continue

            accounts_data.append({
                'code': code,
                'name': name,
                'account_type': account_type,
                'description': description
            })

            if row_num <= 5:
                print(f"  {row_num}: {code} - {name} ({account_type})")

        except Exception as e:
            errors.append(f'Fila {row_num}: {str(e)}')

        row_num += 1

    print(f"\n✓ {len(accounts_data)} cuentas leídas ({len(errors)} errores)")

    if not accounts_data:
        print("✗ No hay datos válidos")
        sys.exit(1)

    # ELIMINAR catálogo anterior
    print(f"\n🗑️  Eliminando catálogo anterior...")
    old_count = session.query(ChartOfAccounts).count()
    session.query(ChartOfAccounts).delete()
    session.commit()
    print(f"  ✓ {old_count} cuentas eliminadas")

    # IMPORTAR nuevos datos
    print(f"\n➕ Importando {len(accounts_data)} cuentas nuevas...")
    for idx, account_data in enumerate(accounts_data, 1):
        account = ChartOfAccounts(**account_data)
        session.add(account)
        if idx % 50 == 0:
            print(f"  {idx} cuentas...")

    session.commit()
    print(f"  ✓ {len(accounts_data)} cuentas importadas")

    # VINCULAR a categorías
    print(f"\n🔗 Vinculando a categorías...")
    categories = session.query(AssetCategory).all()

    for category in categories:
        # Buscar cuenta de depreciación acumulada
        acum_account = session.query(ChartOfAccounts).filter(
            ChartOfAccounts.name.ilike('%Deprec%Acumulada%')
        ).first()

        # Buscar cuenta de gasto de depreciación
        expense_account = session.query(ChartOfAccounts).filter(
            ChartOfAccounts.name.ilike('%Gasto%Deprec%')
        ).first()

        if acum_account:
            category.accumulated_depreciation_account = acum_account.code

        if expense_account:
            category.depreciation_expense_account = expense_account.code

    session.commit()

    print(f"  ✓ {len(categories)} categorías vinculadas")

    print("\n" + "=" * 80)
    print("✅ IMPORTACIÓN COMPLETADA")
    print("=" * 80)
    print(f"\nResumen:")
    print(f"  Cuentas eliminadas: {old_count}")
    print(f"  Cuentas importadas: {len(accounts_data)}")
    print(f"  Categorías vinculadas: {len(categories)}")
    if errors:
        print(f"  Errores: {len(errors)}")
        for err in errors[:5]:
            print(f"    - {err}")
        if len(errors) > 5:
            print(f"    ... y {len(errors) - 5} más")

    session.close()

except Exception as e:
    print(f"\n✗ ERROR: {str(e)}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
