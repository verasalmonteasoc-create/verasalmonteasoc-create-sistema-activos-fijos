#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para importar catálogo de cuentas desde Excel
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

from backend.app import app
from backend.models import db, ChartOfAccounts, AssetCategory

def import_catalog():
    """Importar catálogo de cuentas desde Excel"""
    try:
        print("=" * 80)
        print("IMPORTAR CATÁLOGO DE CUENTAS")
        print("=" * 80)

        # Cargar archivo Excel
        wb = load_workbook('Catalogo_Cuentas.xlsx')
        ws = wb.active

        print(f"\n✓ Archivo cargado: {wb.sheetnames[0]}")
        print(f"  Filas totales: {ws.max_row}")

        # Ver estructura
        print(f"\nEstructura de columnas:")
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        for i, header in enumerate(header_row, 1):
            print(f"  {i}. {header}")

        # Leer datos
        print(f"\nPrimeras 5 cuentas:")
        print("-" * 100)

        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row[0] is None:  # Fin de datos
                break
            data.append(row)
            if row_idx <= 6:
                print(f"Row {row_idx}: {row}")

        print(f"\nTotal de cuentas a importar: {len(data)}")

        # Confirmación
        response = input("\n¿Reemplazar el catálogo anterior? (s/n): ").strip().lower()
        if response != 's':
            print("Cancelado.")
            return

        # Procesar dentro del contexto de la app
        with app.app_context():
            # Eliminar cuentas anteriores
            print("\n🗑️  Eliminando cuentas anteriores...")
            ChartOfAccounts.query.delete()
            db.session.commit()
            print(f"  ✓ {len(data)} cuentas eliminadas")

            # Insertar nuevas cuentas
            print(f"\n➕ Importando {len(data)} cuentas nuevas...")

            for idx, row in enumerate(data, 1):
                code = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if row[1] else ""
                account_type = str(row[2]).strip() if row[2] else "Activo"
                description = str(row[3]).strip() if row[3] else ""

                account = ChartOfAccounts(
                    code=code,
                    name=name,
                    account_type=account_type,
                    description=description
                )
                db.session.add(account)

                if idx % 50 == 0:
                    print(f"  {idx} cuentas procesadas...")

            db.session.commit()
            print(f"  ✓ {len(data)} cuentas importadas exitosamente")

            # Vincular cuentas a categorías
            print(f"\n🔗 Vinculando cuentas a categorías...")
            categories = AssetCategory.query.all()

            # Mapeos automáticos
            mappings = {
                'Deprec. Acumulada': {'accumulated': True},
                'Gasto Depreciación': {'expense': True},
            }

            for category in categories:
                print(f"\n  Categoría: {category.name}")

                # Buscar cuentas de depreciación acumulada
                acum_accounts = ChartOfAccounts.query.filter(
                    ChartOfAccounts.name.ilike('%Deprec%Acumulada%')
                ).all()

                # Buscar cuentas de gasto de depreciación
                expense_accounts = ChartOfAccounts.query.filter(
                    ChartOfAccounts.name.ilike('%Gasto%Deprec%')
                ).all()

                if acum_accounts:
                    category.accumulated_depreciation_account = acum_accounts[0].code
                    print(f"    ✓ Cuenta acumulada: {acum_accounts[0].code}")

                if expense_accounts:
                    category.depreciation_expense_account = expense_accounts[0].code
                    print(f"    ✓ Cuenta gasto: {expense_accounts[0].code}")

            db.session.commit()
            print(f"\n  ✓ Vinculaciones completadas")

        print("\n" + "=" * 80)
        print("✅ IMPORTACIÓN COMPLETADA EXITOSAMENTE")
        print("=" * 80)

    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()

if __name__ == '__main__':
    import_catalog()
