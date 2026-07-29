"""
Ciclo de vida del activo (Fase 2–4):
  - Baja / retiro con cálculo de ganancia o pérdida y su asiento
  - Traslado entre departamentos / localidades (con historial)
  - Revaluación / deterioro
  - Reverso (anulación) de asientos contables
  - Cierre y apertura de períodos contables
  - Conteo físico (inventario) con escaneo de QR
  - Historial de movimientos del activo
"""
from flask import Blueprint, request, jsonify
from flask_login import current_user
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from backend.models import (
    db, Asset, AssetCategory, Department, Location, ChartOfAccounts,
    JournalEntry, JournalEntryLine, AssetMovement, FiscalPeriodLock,
    InventorySession
)

lifecycle_bp = Blueprint('lifecycle', __name__, url_prefix='/api/lifecycle')
_CENT = Decimal('0.01')


# ─────────────────────────── helpers ───────────────────────────
def _require_admin():
    if not current_user.is_authenticated or not current_user.is_admin():
        return jsonify({'success': False, 'message': 'Solo un administrador puede realizar esta operación'}), 403
    return None


def _account_name(code):
    if not code:
        return ''
    acc = ChartOfAccounts.query.filter_by(code=code).first()
    return acc.name if acc else code


def _create_journal(entry_date, description, entry_type, lines, asset_id=None, reference=None):
    """Crea un asiento balanceado. `lines` = lista de dicts con
    {account, debit, credit, desc}. Valida débito == crédito. Devuelve el asiento."""
    total_debit = sum((Decimal(str(l['debit'])) for l in lines), Decimal('0'))
    total_credit = sum((Decimal(str(l['credit'])) for l in lines), Decimal('0'))
    if total_debit.quantize(_CENT) != total_credit.quantize(_CENT):
        raise ValueError(f'Asiento descuadrado: débito {total_debit} ≠ crédito {total_credit}')

    entry = JournalEntry(
        reference=reference,
        entry_date=entry_date,
        description=description,
        amount=total_debit,
        asset_id=asset_id,
        year=entry_date.year,
        month=entry_date.month,
        entry_type=entry_type,
        status='posted',
        total_debit=total_debit,
        total_credit=total_credit,
    )
    db.session.add(entry)
    db.session.flush()
    for l in lines:
        db.session.add(JournalEntryLine(
            journal_entry_id=entry.id,
            account_code=l['account'],
            account_name=_account_name(l['account']),
            description=l.get('desc', description),
            debit_amount=Decimal(str(l['debit'])),
            credit_amount=Decimal(str(l['credit'])),
        ))
    return entry


def _parse_date(s, default=None):
    if not s:
        return default or datetime.utcnow().date()
    if isinstance(s, datetime):
        return s.date()
    return datetime.strptime(str(s).split('T')[0], '%Y-%m-%d').date()


# ─────────────────────────── BAJA / RETIRO ───────────────────────────
@lifecycle_bp.route('/assets/<int:asset_id>/retire', methods=['POST'])
def retire_asset(asset_id):
    """Dar de baja un activo (venta o desecho) con ganancia/pérdida y asiento."""
    err = _require_admin()
    if err:
        return err

    asset = Asset.query.get_or_404(asset_id)
    if asset.status == 'retired':
        return jsonify({'success': False, 'message': 'El activo ya está dado de baja'}), 400

    data = request.get_json() or {}
    disposal_date = _parse_date(data.get('disposal_date'))
    disposal_amount = Decimal(str(data.get('disposal_amount') or 0)).quantize(_CENT)
    reason = (data.get('reason') or 'Baja de activo').strip()
    cash_account = data.get('cash_account')  # cuenta de banco/caja/CxC si hubo cobro

    if FiscalPeriodLock.is_period_closed(disposal_date.year, disposal_date.month):
        return jsonify({'success': False, 'message': f'El período {disposal_date.month}/{disposal_date.year} está cerrado.'}), 409

    cat = asset.category
    if not cat.asset_account or not cat.accumulated_depreciation_account:
        return jsonify({'success': False, 'message': f'La categoría "{cat.name}" no tiene cuentas de activo / depreciación acumulada configuradas.'}), 400

    cost = Decimal(str(asset.acquisition_cost)).quantize(_CENT)
    accumulated = Decimal(str(asset.get_accumulated_depreciation())).quantize(_CENT)
    nbv = (cost - accumulated).quantize(_CENT)
    gain_loss = (disposal_amount - nbv).quantize(_CENT)  # + ganancia / - pérdida

    # Construir el asiento de baja
    lines = []
    # Quitar la depreciación acumulada (débito)
    if accumulated > 0:
        lines.append({'account': cat.accumulated_depreciation_account, 'debit': accumulated, 'credit': 0,
                      'desc': f'Baja acum. deprec. {asset.code}'})
    # Cobro recibido (débito a banco/caja)
    if disposal_amount > 0:
        if not cash_account:
            return jsonify({'success': False, 'message': 'Indica la cuenta de banco/caja para registrar el cobro de la venta.'}), 400
        lines.append({'account': cash_account, 'debit': disposal_amount, 'credit': 0,
                      'desc': f'Cobro venta {asset.code}'})
    # Pérdida (débito)
    if gain_loss < 0:
        if not cat.gain_loss_account:
            return jsonify({'success': False, 'message': f'La categoría "{cat.name}" no tiene cuenta de ganancia/pérdida configurada.'}), 400
        lines.append({'account': cat.gain_loss_account, 'debit': abs(gain_loss), 'credit': 0,
                      'desc': f'Pérdida en baja {asset.code}'})
    # Dar de baja el activo (crédito por el costo)
    lines.append({'account': cat.asset_account, 'debit': 0, 'credit': cost,
                  'desc': f'Baja costo activo {asset.code}'})
    # Ganancia (crédito)
    if gain_loss > 0:
        if not cat.gain_loss_account:
            return jsonify({'success': False, 'message': f'La categoría "{cat.name}" no tiene cuenta de ganancia/pérdida configurada.'}), 400
        lines.append({'account': cat.gain_loss_account, 'debit': 0, 'credit': gain_loss,
                      'desc': f'Ganancia en baja {asset.code}'})

    try:
        entry = _create_journal(disposal_date, f'Baja de activo {asset.code} - {reason}',
                                'disposal', lines, asset_id=asset.id,
                                reference=f'BAJA-{asset.code}')
        asset.status = 'retired'
        asset.disposal_date = disposal_date
        asset.disposal_amount = disposal_amount
        asset.disposal_reason = reason
        asset.disposal_gain_loss = gain_loss
        db.session.add(AssetMovement(
            asset_id=asset.id, movement_type='retirement', movement_date=disposal_date,
            from_value=f'Valor en libros: {nbv}', to_value=reason,
            amount=disposal_amount, journal_entry_id=entry.id,
            notes=f'G/P: {gain_loss}', created_by=current_user.id))
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Activo {asset.code} dado de baja',
            'cost': float(cost), 'accumulated': float(accumulated),
            'net_book_value': float(nbv), 'disposal_amount': float(disposal_amount),
            'gain_loss': float(gain_loss), 'journal_entry_id': entry.id
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400


# ─────────────────────────── TRASLADO ───────────────────────────
@lifecycle_bp.route('/assets/<int:asset_id>/transfer', methods=['POST'])
def transfer_asset(asset_id):
    """Trasladar un activo a otro departamento y/o localidad, con historial."""
    err = _require_admin()
    if err:
        return err
    asset = Asset.query.get_or_404(asset_id)
    data = request.get_json() or {}
    move_date = _parse_date(data.get('date'))

    from_parts, to_parts = [], []
    if 'to_department_id' in data:
        old = asset.department_obj.name if asset.department_obj else '—'
        new_dept = Department.query.get(data['to_department_id']) if data['to_department_id'] else None
        asset.department_id = new_dept.id if new_dept else None
        from_parts.append(f'Depto: {old}')
        to_parts.append(f'Depto: {new_dept.name if new_dept else "—"}')
    if 'to_location_id' in data:
        old = asset.location_obj.name if asset.location_obj else '—'
        new_loc = Location.query.get(data['to_location_id']) if data['to_location_id'] else None
        asset.location_id = new_loc.id if new_loc else None
        from_parts.append(f'Local: {old}')
        to_parts.append(f'Local: {new_loc.name if new_loc else "—"}')
    if data.get('to_user'):
        from_parts.append(f'Usuario: {asset.asset_user or "—"}')
        asset.asset_user = data['to_user']
        to_parts.append(f'Usuario: {data["to_user"]}')

    if not to_parts:
        return jsonify({'success': False, 'message': 'No se indicó ningún cambio de traslado.'}), 400

    db.session.add(AssetMovement(
        asset_id=asset.id, movement_type='transfer', movement_date=move_date,
        from_value=' · '.join(from_parts), to_value=' · '.join(to_parts),
        notes=data.get('notes'), created_by=current_user.id))
    db.session.commit()
    return jsonify({'success': True, 'message': f'Activo {asset.code} trasladado'}), 200


# ─────────────────────────── REVALUACIÓN / DETERIORO ───────────────────────────
@lifecycle_bp.route('/assets/<int:asset_id>/revalue', methods=['POST'])
def revalue_asset(asset_id):
    """Ajuste de valor (revaluación al alza o deterioro a la baja) con asiento."""
    err = _require_admin()
    if err:
        return err
    asset = Asset.query.get_or_404(asset_id)
    data = request.get_json() or {}
    adjustment = Decimal(str(data.get('adjustment') or 0)).quantize(_CENT)  # + revalúa / - deteriora
    move_date = _parse_date(data.get('date'))
    reason = (data.get('reason') or 'Ajuste de valor').strip()

    if adjustment == 0:
        return jsonify({'success': False, 'message': 'El monto del ajuste no puede ser cero.'}), 400
    if FiscalPeriodLock.is_period_closed(move_date.year, move_date.month):
        return jsonify({'success': False, 'message': f'El período {move_date.month}/{move_date.year} está cerrado.'}), 409

    cat = asset.category
    if not cat.asset_account or not cat.gain_loss_account:
        return jsonify({'success': False, 'message': f'La categoría "{cat.name}" necesita cuenta de activo y de ganancia/pérdida.'}), 400

    if adjustment > 0:   # revaluación al alza
        lines = [{'account': cat.asset_account, 'debit': adjustment, 'credit': 0, 'desc': f'Revaluación {asset.code}'},
                 {'account': cat.gain_loss_account, 'debit': 0, 'credit': adjustment, 'desc': f'Superávit revaluación {asset.code}'}]
    else:                # deterioro
        lines = [{'account': cat.gain_loss_account, 'debit': abs(adjustment), 'credit': 0, 'desc': f'Deterioro {asset.code}'},
                 {'account': cat.asset_account, 'debit': 0, 'credit': abs(adjustment), 'desc': f'Deterioro {asset.code}'}]

    try:
        old_cost = Decimal(str(asset.acquisition_cost))
        entry = _create_journal(move_date, f'Ajuste de valor {asset.code} - {reason}', 'revaluation',
                                lines, asset_id=asset.id, reference=f'REV-{asset.code}')
        asset.acquisition_cost = (old_cost + adjustment).quantize(_CENT)
        db.session.add(AssetMovement(
            asset_id=asset.id, movement_type='revaluation', movement_date=move_date,
            from_value=f'Costo: {old_cost}', to_value=f'Costo: {asset.acquisition_cost}',
            amount=adjustment, journal_entry_id=entry.id, notes=reason, created_by=current_user.id))
        db.session.commit()
        return jsonify({'success': True, 'message': f'Valor de {asset.code} ajustado',
                        'new_cost': float(asset.acquisition_cost), 'journal_entry_id': entry.id}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400


# ─────────────────────────── HISTORIAL DE MOVIMIENTOS ───────────────────────────
@lifecycle_bp.route('/assets/<int:asset_id>/movements', methods=['GET'])
def asset_movements(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    movs = AssetMovement.query.filter_by(asset_id=asset.id).order_by(AssetMovement.movement_date.desc(), AssetMovement.id.desc()).all()
    labels = {'acquisition': 'Alta', 'transfer': 'Traslado', 'revaluation': 'Ajuste de valor', 'retirement': 'Baja'}
    return jsonify({'success': True, 'movements': [{
        'id': m.id, 'type': m.movement_type, 'type_label': labels.get(m.movement_type, m.movement_type),
        'date': m.movement_date.isoformat() if m.movement_date else None,
        'from': m.from_value, 'to': m.to_value,
        'amount': float(m.amount) if m.amount is not None else None,
        'journal_entry_id': m.journal_entry_id, 'notes': m.notes
    } for m in movs]}), 200


# ─────────────────────────── REVERSO DE ASIENTOS ───────────────────────────
@lifecycle_bp.route('/journal/<int:entry_id>/reverse', methods=['POST'])
def reverse_journal(entry_id):
    """Anula un asiento creando su reverso (débito↔crédito). No borra nada."""
    err = _require_admin()
    if err:
        return err
    orig = JournalEntry.query.get_or_404(entry_id)
    if orig.is_reversed or orig.status == 'reversed':
        return jsonify({'success': False, 'message': 'Este asiento ya fue reversado.'}), 400
    if orig.entry_type == 'reversal':
        return jsonify({'success': False, 'message': 'No se puede reversar un asiento que ya es un reverso.'}), 400

    lines = JournalEntryLine.query.filter_by(journal_entry_id=orig.id).all()
    if not lines:
        return jsonify({'success': False, 'message': 'El asiento no tiene líneas para reversar.'}), 400

    try:
        rev = JournalEntry(
            reference=f'REV-{orig.reference or orig.id}',
            entry_date=datetime.utcnow().date(),
            description=f'REVERSO de: {orig.description}',
            amount=orig.amount,
            asset_id=orig.asset_id,
            year=datetime.utcnow().year,
            month=datetime.utcnow().month,
            entry_type='reversal',
            status='posted',
            total_debit=orig.total_credit,
            total_credit=orig.total_debit,
            reversal_of_id=orig.id,
        )
        db.session.add(rev)
        db.session.flush()
        for l in lines:
            db.session.add(JournalEntryLine(
                journal_entry_id=rev.id, account_code=l.account_code, account_name=l.account_name,
                description=f'Reverso: {l.description or ""}',
                debit_amount=l.credit_amount, credit_amount=l.debit_amount))  # invertido

        orig.is_reversed = True
        orig.status = 'reversed'

        # Si era un asiento de depreciación, liberar el período (borrar sus registros)
        note = ''
        if orig.entry_type == 'depreciation' and orig.year and orig.month:
            from backend.models import DepreciationRecord
            deleted = DepreciationRecord.query.filter_by(year=orig.year, month=orig.month).delete(synchronize_session=False)
            note = f' Se liberaron {deleted} registros de depreciación de {orig.month}/{orig.year} (puede reprocesarse).'

        db.session.commit()
        return jsonify({'success': True, 'message': f'Asiento reversado.{note}', 'reversal_id': rev.id}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400


# ─────────────────────────── CIERRE DE PERÍODOS ───────────────────────────
@lifecycle_bp.route('/periods', methods=['GET'])
def list_periods():
    locks = FiscalPeriodLock.query.order_by(FiscalPeriodLock.year.desc(), FiscalPeriodLock.month.desc()).all()
    return jsonify({'success': True, 'periods': [{
        'year': l.year, 'month': l.month, 'is_closed': l.is_closed,
        'closed_at': l.closed_at.isoformat() if l.closed_at else None
    } for l in locks]}), 200


@lifecycle_bp.route('/periods/close', methods=['POST'])
def close_period():
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    year, month = data.get('year'), data.get('month')
    if not year or not month:
        return jsonify({'success': False, 'message': 'Año y mes requeridos'}), 400
    lock = FiscalPeriodLock.query.filter_by(year=year, month=month).first()
    if lock:
        lock.is_closed = True
        lock.closed_at = datetime.utcnow()
        lock.closed_by = current_user.id
    else:
        db.session.add(FiscalPeriodLock(year=year, month=month, is_closed=True, closed_by=current_user.id))
    db.session.commit()
    return jsonify({'success': True, 'message': f'Período {month}/{year} cerrado'}), 200


@lifecycle_bp.route('/periods/open', methods=['POST'])
def open_period():
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    year, month = data.get('year'), data.get('month')
    lock = FiscalPeriodLock.query.filter_by(year=year, month=month).first()
    if lock:
        lock.is_closed = False
        db.session.commit()
    return jsonify({'success': True, 'message': f'Período {month}/{year} reabierto'}), 200


# ─────────────────────────── CONTEO FÍSICO (QR) ───────────────────────────
@lifecycle_bp.route('/inventory/sessions', methods=['GET'])
def list_inventory_sessions():
    sessions = InventorySession.query.order_by(InventorySession.started_at.desc()).all()
    return jsonify({'success': True, 'sessions': [{
        'id': s.id, 'name': s.name, 'status': s.status,
        'started_at': s.started_at.isoformat() if s.started_at else None,
        'closed_at': s.closed_at.isoformat() if s.closed_at else None,
        'verified_count': Asset.query.filter_by(last_verified_session_id=s.id).count()
    } for s in sessions]}), 200


@lifecycle_bp.route('/inventory/sessions', methods=['POST'])
def create_inventory_session():
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    name = (data.get('name') or f'Conteo {datetime.utcnow().strftime("%d/%m/%Y")}').strip()
    # Cerrar cualquier sesión abierta previa (solo una abierta a la vez)
    InventorySession.query.filter_by(status='open').update({'status': 'closed', 'closed_at': datetime.utcnow()})
    s = InventorySession(name=name, status='open', started_by=current_user.id)
    db.session.add(s)
    db.session.commit()
    return jsonify({'success': True, 'session': {'id': s.id, 'name': s.name}}), 201


@lifecycle_bp.route('/inventory/sessions/<int:session_id>/close', methods=['POST'])
def close_inventory_session(session_id):
    err = _require_admin()
    if err:
        return err
    s = InventorySession.query.get_or_404(session_id)
    s.status = 'closed'
    s.closed_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'message': 'Sesión de conteo cerrada'}), 200


@lifecycle_bp.route('/inventory/verify', methods=['POST'])
def verify_asset():
    """Marca un activo como verificado (existe físicamente) — lo llama el escaneo del QR."""
    data = request.get_json() or {}
    session = InventorySession.query.filter_by(status='open').order_by(InventorySession.started_at.desc()).first()
    if not session:
        return jsonify({'success': False, 'message': 'No hay una sesión de conteo abierta. Inicia una en Configuración.'}), 400

    asset = None
    if data.get('asset_id'):
        asset = Asset.query.get(data['asset_id'])
    elif data.get('code'):
        asset = Asset.query.filter_by(code=str(data['code']).strip()).first()
    if not asset:
        return jsonify({'success': False, 'message': 'Activo no encontrado'}), 404

    already = asset.last_verified_session_id == session.id
    asset.last_verified_at = datetime.utcnow()
    asset.last_verified_session_id = session.id
    db.session.commit()
    return jsonify({
        'success': True,
        'message': 'Ya estaba verificado en este conteo' if already else 'Activo verificado',
        'already': already,
        'asset': {'code': asset.code, 'description': asset.description,
                  'category': asset.category.name if asset.category else None,
                  'department': asset.department_obj.name if asset.department_obj else None}
    }), 200


@lifecycle_bp.route('/inventory/sessions/<int:session_id>/report', methods=['GET'])
def inventory_report(session_id):
    session = InventorySession.query.get_or_404(session_id)
    active = Asset.query.filter_by(status='active').all()
    found, missing = [], []
    for a in active:
        row = {'code': a.code, 'description': a.description,
               'category': a.category.name if a.category else '-',
               'department': a.department_obj.name if a.department_obj else '-',
               'verified_at': a.last_verified_at.isoformat() if a.last_verified_at else None}
        if a.last_verified_session_id == session.id:
            found.append(row)
        else:
            missing.append(row)
    return jsonify({
        'success': True,
        'session': {'id': session.id, 'name': session.name, 'status': session.status},
        'total': len(active), 'found_count': len(found), 'missing_count': len(missing),
        'found': found, 'missing': missing
    }), 200


# ─────────────────────────── REPORTES (Fase 4) ───────────────────────────
def _xlsx_report(title, headers, rows, filename, totals=None):
    """Genera un Excel con encabezado de empresa (CompanySettings)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from backend.models import CompanySettings
    co = CompanySettings.query.first()
    wb = Workbook(); ws = wb.active; ws.title = 'Reporte'
    r = 1
    if co and (co.legal_name or co.rnc):
        if co.legal_name:
            ws.cell(row=r, column=1, value=co.legal_name).font = Font(bold=True, size=13); r += 1
        if co.rnc:
            ws.cell(row=r, column=1, value=f'RNC: {co.rnc}'); r += 1
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12); r += 1
    ws.cell(row=r, column=1, value=f'Generado: {datetime.utcnow().strftime("%d/%m/%Y")}'); r += 2
    hf = PatternFill(start_color='003D7A', end_color='003D7A', fill_type='solid')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = hf
    r += 1
    for row in rows:
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1
    if totals:
        for c, v in enumerate(totals, 1):
            ws.cell(row=r, column=c, value=v).font = Font(bold=True)
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 24 if i <= 2 else 16
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue(), 200, {
        'Content-Disposition': f'attachment; filename="{filename}.xlsx"',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}


@lifecycle_bp.route('/reports/forecast', methods=['GET'])
def depreciation_forecast():
    """Proyección de depreciación futura (estilo SAP: depreciation forecast)."""
    months = min(int(request.args.get('months', 12)), 60)
    today = datetime.utcnow().date()
    y, m = today.year, today.month
    pending = []
    for a in Asset.query.filter_by(status='active').all():
        if not a.useful_life_years:
            continue
        cost = Decimal(str(a.acquisition_cost))
        base = cost - cost * Decimal(str(a.residual_value_percent or 0)) / Decimal('100')
        rem = base - Decimal(str(a.get_accumulated_depreciation()))
        if rem <= 0:
            continue
        quota = (base / Decimal(a.useful_life_years * 12)).quantize(_CENT, rounding=ROUND_HALF_UP)
        pending.append([rem, quota])
    rows = []
    for _ in range(months):
        m += 1
        if m > 12:
            m = 1; y += 1
        total = Decimal('0'); n = 0
        for it in pending:
            if it[0] <= 0:
                continue
            q = min(it[1], it[0]); it[0] -= q; total += q; n += 1
        rows.append({'period': f'{y}-{m:02d}', 'total': float(total), 'assets': n})
    if request.args.get('format') == 'xlsx':
        return _xlsx_report('Proyección de Depreciación', ['Período', 'Depreciación Proyectada', 'Activos'],
                            [[r['period'], r['total'], r['assets']] for r in rows], 'Proyeccion_Depreciacion')
    return jsonify({'success': True, 'rows': rows}), 200

@lifecycle_bp.route('/reports/movement', methods=['GET'])
def movement_report():
    """Movimiento de activos (roll-forward) por categoría en un rango de fechas:
    adiciones, bajas y depreciación del período."""
    from backend.models import DepreciationRecord
    date_from = _parse_date(request.args.get('from'), datetime(datetime.utcnow().year, 1, 1).date())
    date_to = _parse_date(request.args.get('to'))

    cats = {}
    def _c(cat):
        if cat.id not in cats:
            cats[cat.id] = {'category': cat.name, 'additions': Decimal('0'), 'additions_n': 0,
                            'retirements': Decimal('0'), 'retirements_n': 0, 'depreciation': Decimal('0')}
        return cats[cat.id]

    for a in Asset.query.all():
        if not a.category:
            continue
        if a.acquisition_date and date_from <= a.acquisition_date <= date_to:
            c = _c(a.category); c['additions'] += Decimal(str(a.acquisition_cost)); c['additions_n'] += 1
        if a.disposal_date and date_from <= a.disposal_date <= date_to:
            c = _c(a.category); c['retirements'] += Decimal(str(a.acquisition_cost)); c['retirements_n'] += 1

    for rec in DepreciationRecord.query.all():
        a = Asset.query.get(rec.asset_id)
        if not a or not a.category:
            continue
        try:
            rec_date = datetime(rec.year, rec.month, 1).date()
        except Exception:
            continue
        if date_from <= rec_date <= date_to:
            _c(a.category)['depreciation'] += Decimal(str(rec.depreciation_amount))

    rows = sorted([{'category': v['category'], 'additions': float(v['additions']), 'additions_n': v['additions_n'],
                    'retirements': float(v['retirements']), 'retirements_n': v['retirements_n'],
                    'depreciation': float(v['depreciation'])} for v in cats.values()],
                  key=lambda r: r['category'])
    if request.args.get('format') == 'xlsx':
        return _xlsx_report(f'Movimiento de Activos {date_from} a {date_to}',
                            ['Categoría', 'Adiciones', '# Altas', 'Bajas', '# Bajas', 'Depreciación'],
                            [[r['category'], r['additions'], r['additions_n'], r['retirements'],
                              r['retirements_n'], r['depreciation']] for r in rows], 'Movimiento_Activos')
    return jsonify({'success': True, 'from': date_from.isoformat(), 'to': date_to.isoformat(),
                    'rows': rows}), 200


@lifecycle_bp.route('/reports/register', methods=['GET'])
def asset_register():
    """Libro / registro de activos fijos: costo, deprec. acumulada y valor en libros."""
    include_retired = request.args.get('include_retired') == '1'
    q = Asset.query if include_retired else Asset.query.filter_by(status='active')
    rows, tc, ta, tn = [], Decimal('0'), Decimal('0'), Decimal('0')
    for a in q.order_by(Asset.code).all():
        cost = Decimal(str(a.acquisition_cost))
        accum = Decimal(str(a.get_accumulated_depreciation()))
        nbv = cost - accum
        tc += cost; ta += accum; tn += nbv
        rows.append({'code': a.code, 'description': a.description,
                     'category': a.category.name if a.category else '-',
                     'department': a.department_obj.name if a.department_obj else '-',
                     'acquisition_date': a.acquisition_date.isoformat() if a.acquisition_date else None,
                     'status': a.status, 'cost': float(cost), 'accumulated': float(accum), 'nbv': float(nbv)})
    if request.args.get('format') == 'xlsx':
        return _xlsx_report('Libro de Activos Fijos',
                            ['Código', 'Descripción', 'Categoría', 'Departamento', 'F. Adq.', 'Costo', 'Dep. Acum.', 'Valor en Libros'],
                            [[r['code'], r['description'], r['category'], r['department'], r['acquisition_date'],
                              r['cost'], r['accumulated'], r['nbv']] for r in rows],
                            'Libro_Activos_Fijos',
                            totals=['TOTAL', '', '', '', '', float(tc), float(ta), float(tn)])
    return jsonify({'success': True, 'rows': rows,
                    'totals': {'cost': float(tc), 'accumulated': float(ta), 'nbv': float(tn), 'count': len(rows)}}), 200


@lifecycle_bp.route('/reports/tax-vs-book', methods=['GET'])
def tax_vs_book():
    """Cédula fiscal vs contable: compara la depreciación contable (tasa NIIF de
    la categoría) con la fiscal (tax_depreciation_rate) por categoría."""
    rows = []
    t_book, t_tax = Decimal('0'), Decimal('0')
    for cat in AssetCategory.query.order_by(AssetCategory.name).all():
        assets = cat.assets.filter_by(status='active').all()
        if not assets:
            continue
        base = sum(Decimal(str(a.acquisition_cost)) for a in assets)
        book_rate = Decimal(str(cat.depreciation_rate or 0))
        tax_rate = Decimal(str(cat.tax_depreciation_rate)) if cat.tax_depreciation_rate is not None else book_rate
        book_annual = (base * book_rate / Decimal('100')).quantize(_CENT)
        tax_annual = (base * tax_rate / Decimal('100')).quantize(_CENT)
        t_book += book_annual; t_tax += tax_annual
        rows.append({'category': cat.name, 'assets': len(assets), 'base': float(base),
                     'book_rate': float(book_rate), 'tax_rate': float(tax_rate),
                     'book_annual': float(book_annual), 'tax_annual': float(tax_annual),
                     'difference': float(book_annual - tax_annual)})
    if request.args.get('format') == 'xlsx':
        return _xlsx_report('Depreciación Anual: Fiscal vs Contable',
                            ['Categoría', 'Base', 'Tasa NIIF %', 'Tasa Fiscal %', 'Contable', 'Fiscal', 'Diferencia'],
                            [[r['category'], r['base'], r['book_rate'], r['tax_rate'],
                              r['book_annual'], r['tax_annual'], r['difference']] for r in rows],
                            'Fiscal_vs_Contable',
                            totals=['TOTAL', '', '', '', float(t_book), float(t_tax), float(t_book - t_tax)])
    return jsonify({'success': True, 'rows': rows,
                    'totals': {'book': float(t_book), 'tax': float(t_tax), 'difference': float(t_book - t_tax)}}), 200
