"""
Rutas de Reportes - SAP-style Fixed Asset Management
"""
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required
from datetime import datetime, timedelta
from decimal import Decimal
from backend.models import db, Asset, AssetCategory, DepreciationRecord, AuditLog, JournalEntry, JournalEntryLine
from io import BytesIO
import csv
import logging

logger = logging.getLogger(__name__)

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


@reports_bp.route('/assets-summary', methods=['GET'])
def get_assets_summary():
    """Resumen general de activos"""
    try:
        total_assets = Asset.query.count()
        active_assets = Asset.query.filter_by(status='active').count()
        inactive_assets = Asset.query.filter_by(status='inactive').count()
        retired_assets = Asset.query.filter_by(status='retired').count()

        total_cost = db.session.query(db.func.sum(Asset.acquisition_cost)).scalar() or 0
        total_depreciation = 0
        total_net_value = 0

        for asset in Asset.query.all():
            total_depreciation += asset.get_accumulated_depreciation()
            total_net_value += asset.get_net_book_value()

        return jsonify({
            'success': True,
            'summary': {
                'total_assets': total_assets,
                'active_assets': active_assets,
                'inactive_assets': inactive_assets,
                'retired_assets': retired_assets,
                'total_acquisition_cost': float(total_cost),
                'total_accumulated_depreciation': total_depreciation,
                'total_net_book_value': total_net_value
            }
        }), 200
    except Exception as e:
        logger.error(f"Error en assets-summary: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400


@reports_bp.route('/by-category', methods=['GET'])
def get_by_category():
    """Reporte de activos por categoría"""
    try:
        categories = AssetCategory.query.all()
        report = []

        for cat in categories:
            assets = Asset.query.filter_by(category_id=cat.id).all()
            total_cost = sum(float(a.acquisition_cost) for a in assets)
            total_depreciation = sum(a.get_accumulated_depreciation() for a in assets)
            total_net_value = sum(a.get_net_book_value() for a in assets)

            report.append({
                'category': cat.name,
                'depreciation_rate': float(cat.depreciation_rate),
                'asset_count': len(assets),
                'total_acquisition_cost': total_cost,
                'total_accumulated_depreciation': total_depreciation,
                'total_net_book_value': total_net_value
            })

        return jsonify({
            'success': True,
            'report': report
        }), 200
    except Exception as e:
        logger.error(f"Error en by-category: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400


@reports_bp.route('/depreciation', methods=['GET'])
def get_depreciation_report():
    """Reporte de depreciación por período"""
    try:
        year = request.args.get('year', datetime.utcnow().year, type=int)
        month = request.args.get('month', type=int)

        query = DepreciationRecord.query.filter_by(year=year)
        if month:
            query = query.filter_by(month=month)

        records = query.order_by(
            DepreciationRecord.asset_id,
            DepreciationRecord.month
        ).all()

        total_depreciation = sum(float(r.depreciation_amount) for r in records)

        return jsonify({
            'success': True,
            'period': {
                'year': year,
                'month': month
            },
            'total_depreciation': total_depreciation,
            'records': [{
                'asset_code': r.asset.code,
                'asset_description': r.asset.description,
                'category': r.asset.category.name,
                'month_year': r.get_display_month(),
                'depreciation_amount': float(r.depreciation_amount),
                'accumulated_depreciation': float(r.accumulated_depreciation),
                'net_book_value': float(r.net_book_value)
            } for r in records],
            'record_count': len(records)
        }), 200
    except Exception as e:
        logger.error(f"Error en depreciation: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400


@reports_bp.route('/depreciation-detail-excel', methods=['GET'])
def get_depreciation_detail_excel():
    """Reporte detallado de depreciación por activo (Excel)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        year = request.args.get('year', datetime.utcnow().year, type=int)
        month = request.args.get('month', type=int)

        query = DepreciationRecord.query.filter_by(year=year)
        if month:
            query = query.filter_by(month=month)

        records = query.order_by(DepreciationRecord.asset_id).all()

        if not records:
            return jsonify({'success': False, 'message': 'No hay datos de depreciación para el período'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Depreciación Detallada"

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        currency_format = '_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-'

        headers = ['Código', 'Descripción', 'Categoría', 'Mes/Año', 'Costo Adquisición',
                   'Depreciación Mensual', 'Depreciación Acumulada', 'Valor Neto en Libros']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        row = 2
        for record in records:
            ws.cell(row=row, column=1).value = record.asset.code
            ws.cell(row=row, column=2).value = record.asset.description
            ws.cell(row=row, column=3).value = record.asset.category.name
            ws.cell(row=row, column=4).value = record.get_display_month()
            ws.cell(row=row, column=5).value = float(record.asset.acquisition_cost)
            ws.cell(row=row, column=6).value = float(record.depreciation_amount)
            ws.cell(row=row, column=7).value = float(record.accumulated_depreciation)
            ws.cell(row=row, column=8).value = float(record.net_book_value)

            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col >= 5:
                    cell.number_format = currency_format
                cell.alignment = Alignment(horizontal='right' if col >= 5 else 'left')

            row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        for col in ['E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'reporte_depreciacion_detallada_{year}'
        if month:
            filename += f'{month:02d}'
        filename += '.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"Error generando depreciation-detail-excel: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400


@reports_bp.route('/reconciliation-excel', methods=['GET'])
def get_reconciliation_excel():
    """Reporte de reconciliación activos vs contabilidad"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Reconciliación"

        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        currency_format = '_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-'

        headers = ['Categoría', 'Activos en Sistema', 'Costo Total Sistema', 'Depreciación Sistema',
                   'Valor Neto Sistema', 'Varianza']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        categories = AssetCategory.query.all()
        row = 2

        total_assets = 0
        total_cost = 0
        total_depreciation = 0
        total_net = 0

        for cat in categories:
            assets = Asset.query.filter_by(category_id=cat.id).all()

            cat_total_cost = sum(float(a.acquisition_cost) for a in assets)
            cat_total_depreciation = sum(a.get_accumulated_depreciation() for a in assets)
            cat_total_net = sum(a.get_net_book_value() for a in assets)

            ws.cell(row=row, column=1).value = cat.name
            ws.cell(row=row, column=2).value = len(assets)
            ws.cell(row=row, column=3).value = cat_total_cost
            ws.cell(row=row, column=4).value = cat_total_depreciation
            ws.cell(row=row, column=5).value = cat_total_net
            ws.cell(row=row, column=6).value = 0

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col >= 3:
                    cell.number_format = currency_format
                    cell.alignment = Alignment(horizontal='right')

            total_assets += len(assets)
            total_cost += cat_total_cost
            total_depreciation += cat_total_depreciation
            total_net += cat_total_net
            row += 1

        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=2).value = total_assets
        ws.cell(row=row, column=3).value = total_cost
        ws.cell(row=row, column=4).value = total_depreciation
        ws.cell(row=row, column=5).value = total_net

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.border = border
            if col >= 3:
                cell.number_format = currency_format

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 22

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_reconciliacion_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Error generando reconciliation-excel: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400


@reports_bp.route('/audit-trail-excel', methods=['GET'])
def get_audit_trail_excel():
    """Reporte de pista de auditoría"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side

        days = request.args.get('days', 30, type=int)
        cutoff_date = datetime.utcnow() - timedelta(days=days)

        logs = AuditLog.query.filter(AuditLog.timestamp >= cutoff_date).order_by(AuditLog.timestamp.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Pista de Auditoría"

        header_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Fecha/Hora', 'Usuario', 'Entidad', 'ID', 'Acción', 'Descripción', 'Dirección IP']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        row = 2
        for log in logs:
            user_name = log.user.username if log.user else "Sistema"

            ws.cell(row=row, column=1).value = log.timestamp
            ws.cell(row=row, column=2).value = user_name
            ws.cell(row=row, column=3).value = log.entity_type
            ws.cell(row=row, column=4).value = log.entity_id or ""
            ws.cell(row=row, column=5).value = log.action
            ws.cell(row=row, column=6).value = log.description or ""
            ws.cell(row=row, column=7).value = log.ip_address or ""

            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col == 1:
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'

            row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_pista_auditoria_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Error generando audit-trail-excel: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400
