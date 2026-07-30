"""
Módulo de Inventario Físico de Activos Fijos.

Flujo completo (estilo SAP MI/Inventario físico, operable desde el celular):
  1. Se abre una sesión de conteo (opcionalmente con alcance: un departamento
     o una categoría, para repartir el trabajo).
  2. En el campo se escanea el QR de cada activo con la cámara del celular
     (o se escribe su código). Cada escaneo registra una línea de conteo con
     dónde y en qué estado se encontró el activo.
  3. El sistema detecta automáticamente DIFERENCIAS: activo en otro
     departamento/localidad del que dice el maestro, o en otra condición.
  4. Se registran también los activos hallados que NO están en el sistema.
  5. Al cerrar: reporte de contados / faltantes / diferencias / no registrados
     (en pantalla y en Excel), y opción de APLICAR los hallazgos al maestro.
"""
from flask import Blueprint, request, jsonify
from flask_login import current_user
from datetime import datetime
from urllib.parse import urlparse, parse_qs
from backend.models import (
    db, Asset, AssetCategory, Department, Location, User,
    InventorySession, InventoryCount, InventoryUnregistered, AssetMovement
)

inventory_bp = Blueprint('inventory', __name__, url_prefix='/api/inventory')

CONDITION_LABELS = {'good': 'Bueno', 'fair': 'Regular', 'poor': 'Malo', 'retired': 'Retirado'}


def _require_admin():
    if not current_user.is_authenticated or not current_user.is_admin():
        return jsonify({'success': False, 'message': 'Solo un administrador puede hacer esto'}), 403
    return None


def resolve_scanned_code(raw):
    """Convierte lo que devuelve el escáner en un activo.

    Acepta:
      - El código del activo: "VEH-001"
      - La URL del QR: "https://af.aplicacionesrd.com/?asset_id=995"
      - El chasis/VIN o el número de serie
    """
    if not raw:
        return None
    text = str(raw).strip()

    # 1) URL con ?asset_id=
    if '://' in text or text.startswith('/?'):
        try:
            qs = parse_qs(urlparse(text).query)
            if 'asset_id' in qs:
                a = Asset.query.get(int(qs['asset_id'][0]))
                if a:
                    return a
        except (ValueError, TypeError):
            pass

    # 2) Sólo dígitos → id
    if text.isdigit():
        a = Asset.query.get(int(text))
        if a:
            return a

    # 3) Código exacto (case-insensitive)
    a = Asset.query.filter(db.func.upper(Asset.code) == text.upper()).first()
    if a:
        return a

    # 4) Chasis / serie
    return Asset.query.filter(
        db.or_(db.func.upper(Asset.chassis) == text.upper(),
               db.func.upper(Asset.serial_number) == text.upper())
    ).first()


def _scope_query(session):
    """Activos que entran en el alcance de la sesión."""
    q = Asset.query.filter(Asset.status == 'active')
    if session.scope_department_id:
        q = q.filter(Asset.department_id == session.scope_department_id)
    if session.scope_category_id:
        q = q.filter(Asset.category_id == session.scope_category_id)
    return q


def _session_stats(session):
    total = _scope_query(session).count()
    counted = session.counts.count()
    discrepancies = session.counts.filter_by(has_discrepancy=True).count()
    extras = InventoryUnregistered.query.filter_by(session_id=session.id).count()
    return {
        'total': total, 'counted': counted,
        'missing': max(0, total - counted),
        'discrepancies': discrepancies, 'unregistered': extras,
        'progress': round(counted * 100.0 / total, 1) if total else 0.0
    }


def _serialize_session(s):
    d = {
        'id': s.id, 'name': s.name, 'status': s.status,
        'started_at': s.started_at.isoformat() if s.started_at else None,
        'closed_at': s.closed_at.isoformat() if s.closed_at else None,
        'applied_at': s.applied_at.isoformat() if s.applied_at else None,
        'scope_department_id': s.scope_department_id,
        'scope_category_id': s.scope_category_id,
        'notes': s.notes,
    }
    dept = Department.query.get(s.scope_department_id) if s.scope_department_id else None
    cat = AssetCategory.query.get(s.scope_category_id) if s.scope_category_id else None
    scope = []
    if dept:
        scope.append(f'Depto: {dept.name}')
    if cat:
        scope.append(f'Categoría: {cat.name}')
    d['scope_label'] = ' · '.join(scope) if scope else 'Todos los activos'
    d['stats'] = _session_stats(s)
    return d


# ───────────────────────────── SESIONES ─────────────────────────────
@inventory_bp.route('/sessions', methods=['GET'])
def list_sessions():
    sessions = InventorySession.query.order_by(InventorySession.started_at.desc()).all()
    return jsonify({'success': True, 'sessions': [_serialize_session(s) for s in sessions]}), 200


@inventory_bp.route('/sessions/open', methods=['GET'])
def get_open_session():
    """La sesión abierta actual (la que usa la pantalla de conteo)."""
    s = InventorySession.query.filter_by(status='open').order_by(InventorySession.started_at.desc()).first()
    return jsonify({'success': True, 'session': _serialize_session(s) if s else None}), 200


@inventory_bp.route('/sessions', methods=['POST'])
def create_session():
    err = _require_admin()
    if err:
        return err
    data = request.get_json() or {}
    name = (data.get('name') or f'Conteo {datetime.utcnow().strftime("%d/%m/%Y")}').strip()

    # Solo una sesión abierta a la vez
    InventorySession.query.filter_by(status='open').update(
        {'status': 'closed', 'closed_at': datetime.utcnow()})

    s = InventorySession(
        name=name, status='open', started_by=current_user.id,
        scope_department_id=data.get('scope_department_id') or None,
        scope_category_id=data.get('scope_category_id') or None,
        notes=data.get('notes'))
    db.session.add(s)
    db.session.commit()
    return jsonify({'success': True, 'session': _serialize_session(s)}), 201


@inventory_bp.route('/sessions/<int:session_id>/close', methods=['POST'])
def close_session(session_id):
    err = _require_admin()
    if err:
        return err
    s = InventorySession.query.get_or_404(session_id)
    s.status = 'closed'
    s.closed_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'message': 'Conteo cerrado', 'session': _serialize_session(s)}), 200


@inventory_bp.route('/sessions/<int:session_id>/reopen', methods=['POST'])
def reopen_session(session_id):
    err = _require_admin()
    if err:
        return err
    s = InventorySession.query.get_or_404(session_id)
    InventorySession.query.filter_by(status='open').update(
        {'status': 'closed', 'closed_at': datetime.utcnow()})
    s.status = 'open'
    s.closed_at = None
    db.session.commit()
    return jsonify({'success': True, 'message': 'Conteo reabierto', 'session': _serialize_session(s)}), 200


# ───────────────────────────── ESCANEO / CONTEO ─────────────────────────────
@inventory_bp.route('/lookup', methods=['GET'])
def lookup():
    """Resuelve un código/QR y devuelve la ficha del activo, SIN registrar nada.
    Lo usa el escáner para mostrar el activo antes de confirmar el conteo."""
    asset = resolve_scanned_code(request.args.get('code'))
    if not asset:
        return jsonify({'success': False, 'message': 'No se encontró ningún activo con ese código'}), 404

    session = InventorySession.query.filter_by(status='open').order_by(InventorySession.started_at.desc()).first()
    already = None
    if session:
        c = InventoryCount.query.filter_by(session_id=session.id, asset_id=asset.id).first()
        if c:
            already = {'counted_at': c.counted_at.isoformat(), 'has_discrepancy': c.has_discrepancy}

    return jsonify({'success': True, 'already_counted': already, 'asset': {
        'id': asset.id, 'code': asset.code, 'description': asset.description,
        'category': asset.category.name if asset.category else None,
        'department': asset.department_obj.name if asset.department_obj else None,
        'department_id': asset.department_id,
        'location': asset.location_obj.name if asset.location_obj else None,
        'location_id': asset.location_id,
        'condition': asset.asset_condition,
        'condition_label': CONDITION_LABELS.get(asset.asset_condition, asset.asset_condition),
        'asset_user': asset.asset_user, 'brand': asset.brand, 'model': asset.model,
        'chassis': asset.chassis, 'plate_number': asset.plate_number,
        'status': asset.status,
        'acquisition_cost': float(asset.acquisition_cost),
        'net_book_value': asset.get_net_book_value(),
    }}), 200


@inventory_bp.route('/count', methods=['POST'])
def register_count():
    """Registra (o actualiza) el conteo de un activo en la sesión abierta.
    Detecta diferencias automáticamente contra el maestro."""
    data = request.get_json() or {}

    session = None
    if data.get('session_id'):
        session = InventorySession.query.get(data['session_id'])
    if not session:
        session = InventorySession.query.filter_by(status='open').order_by(InventorySession.started_at.desc()).first()
    if not session:
        return jsonify({'success': False, 'message': 'No hay un conteo abierto. Crea uno primero.'}), 400
    if session.status != 'open':
        return jsonify({'success': False, 'message': 'Ese conteo ya está cerrado.'}), 400

    asset = None
    if data.get('asset_id'):
        asset = Asset.query.get(data['asset_id'])
    if not asset:
        asset = resolve_scanned_code(data.get('code'))
    if not asset:
        return jsonify({'success': False, 'message': 'Activo no encontrado'}), 404

    found_dept_id = data.get('found_department_id') or asset.department_id
    found_loc_id = data.get('found_location_id') or asset.location_id
    found_cond = data.get('found_condition') or asset.asset_condition

    # Detección de diferencias vs el maestro
    diffs = []
    if found_dept_id and asset.department_id and int(found_dept_id) != asset.department_id:
        d_old = asset.department_obj.name if asset.department_obj else '—'
        d_new = Department.query.get(found_dept_id)
        diffs.append(f'Departamento: registrado "{d_old}" → hallado "{d_new.name if d_new else "?"}"')
    if found_loc_id and asset.location_id and int(found_loc_id) != asset.location_id:
        l_old = asset.location_obj.name if asset.location_obj else '—'
        l_new = Location.query.get(found_loc_id)
        diffs.append(f'Localidad: registrada "{l_old}" → hallada "{l_new.name if l_new else "?"}"')
    if found_cond and found_cond != asset.asset_condition:
        diffs.append(f'Condición: registrada "{CONDITION_LABELS.get(asset.asset_condition, asset.asset_condition)}" '
                     f'→ hallada "{CONDITION_LABELS.get(found_cond, found_cond)}"')

    count = InventoryCount.query.filter_by(session_id=session.id, asset_id=asset.id).first()
    is_new = count is None
    if is_new:
        count = InventoryCount(session_id=session.id, asset_id=asset.id)
        db.session.add(count)

    count.counted_at = datetime.utcnow()
    count.counted_by = current_user.id if current_user.is_authenticated else None
    count.method = data.get('method') or 'qr'
    count.found_department_id = found_dept_id
    count.found_location_id = found_loc_id
    count.found_condition = found_cond
    count.observations = data.get('observations')
    count.has_discrepancy = bool(diffs)
    count.discrepancy_notes = ' | '.join(diffs)[:500] if diffs else None

    # Sello rápido en el activo (para "última vez visto")
    asset.last_verified_at = count.counted_at
    asset.last_verified_session_id = session.id

    db.session.commit()

    return jsonify({
        'success': True,
        'is_new': is_new,
        'message': 'Activo contado' if is_new else 'Conteo actualizado',
        'has_discrepancy': count.has_discrepancy,
        'discrepancies': diffs,
        'asset': {'code': asset.code, 'description': asset.description},
        'stats': _session_stats(session)
    }), 200


@inventory_bp.route('/count/<int:count_id>', methods=['DELETE'])
def delete_count(count_id):
    """Deshace un conteo (si se escaneó por error)."""
    err = _require_admin()
    if err:
        return err
    c = InventoryCount.query.get_or_404(count_id)
    session = c.session
    db.session.delete(c)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Conteo eliminado', 'stats': _session_stats(session)}), 200


# ───────────────────────── ACTIVOS NO REGISTRADOS ─────────────────────────
@inventory_bp.route('/sessions/<int:session_id>/unregistered', methods=['POST'])
def add_unregistered(session_id):
    """Registra un activo hallado físicamente que no existe en el sistema."""
    session = InventorySession.query.get_or_404(session_id)
    data = request.get_json() or {}
    desc = (data.get('description') or '').strip()
    if not desc:
        return jsonify({'success': False, 'message': 'La descripción es obligatoria'}), 400
    item = InventoryUnregistered(
        session_id=session.id, description=desc,
        scanned_code=data.get('scanned_code'),
        department_id=data.get('department_id') or None,
        location_id=data.get('location_id') or None,
        observations=data.get('observations'),
        created_by=current_user.id if current_user.is_authenticated else None)
    db.session.add(item)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Registrado como hallazgo no inventariado',
                    'stats': _session_stats(session)}), 201


@inventory_bp.route('/unregistered/<int:item_id>', methods=['DELETE'])
def delete_unregistered(item_id):
    err = _require_admin()
    if err:
        return err
    item = InventoryUnregistered.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Hallazgo eliminado'}), 200


# ───────────────────────────── REPORTE ─────────────────────────────
def _build_report(session):
    counted_rows, missing_rows = [], []
    counts = {c.asset_id: c for c in session.counts.all()}

    for c in session.counts.order_by(InventoryCount.counted_at.desc()).all():
        a = c.asset
        if not a:
            continue
        user = User.query.get(c.counted_by) if c.counted_by else None
        counted_rows.append({
            'count_id': c.id, 'asset_id': a.id, 'code': a.code, 'description': a.description,
            'category': a.category.name if a.category else '-',
            'department': a.department_obj.name if a.department_obj else '-',
            'found_department': c.found_department.name if c.found_department else '-',
            'found_location': c.found_location.name if c.found_location else '-',
            'found_condition': CONDITION_LABELS.get(c.found_condition, c.found_condition or '-'),
            'counted_at': c.counted_at.isoformat() if c.counted_at else None,
            'counted_by': (user.first_name or user.username) if user else '-',
            'method': c.method, 'observations': c.observations or '',
            'has_discrepancy': c.has_discrepancy,
            'discrepancy_notes': c.discrepancy_notes or ''
        })

    for a in _scope_query(session).order_by(Asset.code).all():
        if a.id in counts:
            continue
        missing_rows.append({
            'asset_id': a.id, 'code': a.code, 'description': a.description,
            'category': a.category.name if a.category else '-',
            'department': a.department_obj.name if a.department_obj else '-',
            'location': a.location_obj.name if a.location_obj else '-',
            'asset_user': a.asset_user or '-',
            'cost': float(a.acquisition_cost),
            'nbv': a.get_net_book_value(),
            'last_verified_at': a.last_verified_at.isoformat() if a.last_verified_at else None
        })

    extras = [{
        'id': u.id, 'description': u.description, 'scanned_code': u.scanned_code or '-',
        'department': (Department.query.get(u.department_id).name if u.department_id else '-'),
        'location': (Location.query.get(u.location_id).name if u.location_id else '-'),
        'observations': u.observations or '',
        'created_at': u.created_at.isoformat() if u.created_at else None
    } for u in InventoryUnregistered.query.filter_by(session_id=session.id).all()]

    return counted_rows, missing_rows, extras


@inventory_bp.route('/sessions/<int:session_id>/report', methods=['GET'])
def report(session_id):
    session = InventorySession.query.get_or_404(session_id)
    counted, missing, extras = _build_report(session)
    if request.args.get('format') == 'xlsx':
        return _export_xlsx(session, counted, missing, extras)
    return jsonify({
        'success': True,
        'session': _serialize_session(session),
        'counted': counted,
        'missing': missing,
        'unregistered': extras,
        'discrepancies': [c for c in counted if c['has_discrepancy']]
    }), 200


def _export_xlsx(session, counted, missing, extras):
    """Excel de varias hojas: Resumen · Contados · Faltantes · Diferencias · No registrados."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from backend.models import CompanySettings

    co = CompanySettings.query.first()
    stats = _session_stats(session)
    scope_label = _serialize_session(session)['scope_label']
    hf = PatternFill(start_color='003D7A', end_color='003D7A', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF')
    wb = Workbook()

    def sheet(title, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        r = 1
        if co and co.legal_name:
            ws.cell(row=r, column=1, value=co.legal_name).font = Font(bold=True, size=13); r += 1
            if co.rnc:
                ws.cell(row=r, column=1, value=f'RNC: {co.rnc}'); r += 1
        ws.cell(row=r, column=1, value=f'Inventario Físico — {session.name}').font = Font(bold=True, size=12); r += 1
        ws.cell(row=r, column=1, value=f'{scope_label} · Generado: {datetime.utcnow().strftime("%d/%m/%Y %H:%M")}'); r += 2
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=h); cell.font = hfont; cell.fill = hf
        r += 1
        for row in rows:
            for c, v in enumerate(row, 1):
                ws.cell(row=r, column=c, value=v)
            r += 1
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + i)].width = 30 if i == 2 else 18
        return ws

    # Resumen
    ws = sheet('Resumen', ['Concepto', 'Valor'], [
        ['Conteo', session.name],
        ['Estado', session.status],
        ['Iniciado', session.started_at.strftime('%d/%m/%Y %H:%M') if session.started_at else ''],
        ['Activos en alcance', stats['total']],
        ['Contados (encontrados)', stats['counted']],
        ['Faltantes (no hallados)', stats['missing']],
        ['Con diferencias', stats['discrepancies']],
        ['Hallados sin registrar', stats['unregistered']],
        ['Avance %', stats['progress']],
    ], first=True)

    sheet('Contados', ['Código', 'Descripción', 'Categoría', 'Depto registrado', 'Depto hallado',
                       'Localidad hallada', 'Condición', 'Contado por', 'Fecha', 'Método',
                       'Diferencia', 'Detalle diferencia', 'Observaciones'],
          [[c['code'], c['description'], c['category'], c['department'], c['found_department'],
            c['found_location'], c['found_condition'], c['counted_by'],
            (c['counted_at'] or '')[:16].replace('T', ' '), c['method'],
            'SÍ' if c['has_discrepancy'] else 'No', c['discrepancy_notes'], c['observations']]
           for c in counted])

    sheet('Faltantes', ['Código', 'Descripción', 'Categoría', 'Departamento', 'Localidad',
                        'Usuario', 'Costo', 'Valor en libros', 'Última verificación'],
          [[m['code'], m['description'], m['category'], m['department'], m['location'],
            m['asset_user'], m['cost'], m['nbv'],
            (m['last_verified_at'] or 'Nunca')[:16].replace('T', ' ')] for m in missing])

    sheet('Diferencias', ['Código', 'Descripción', 'Detalle de la diferencia', 'Observaciones'],
          [[c['code'], c['description'], c['discrepancy_notes'], c['observations']]
           for c in counted if c['has_discrepancy']])

    sheet('No registrados', ['Descripción', 'Código escaneado', 'Departamento', 'Localidad', 'Observaciones'],
          [[e['description'], e['scanned_code'], e['department'], e['location'], e['observations']]
           for e in extras])

    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname = f'Inventario_{session.id}_{datetime.utcnow().strftime("%Y%m%d")}'
    return out.getvalue(), 200, {
        'Content-Disposition': f'attachment; filename="{fname}.xlsx"',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}


# ─────────────────────── APLICAR HALLAZGOS AL MAESTRO ───────────────────────
@inventory_bp.route('/sessions/<int:session_id>/apply', methods=['POST'])
def apply_findings(session_id):
    """Actualiza el maestro de activos con lo hallado en el conteo:
    departamento, localidad y condición. Deja traza en el historial del activo."""
    err = _require_admin()
    if err:
        return err
    session = InventorySession.query.get_or_404(session_id)
    applied = 0

    for c in session.counts.filter_by(has_discrepancy=True).all():
        a = c.asset
        if not a:
            continue
        changes = []
        if c.found_department_id and c.found_department_id != a.department_id:
            old = a.department_obj.name if a.department_obj else '—'
            a.department_id = c.found_department_id
            new = a.department_obj.name if a.department_obj else '—'
            changes.append(f'Depto: {old} → {new}')
        if c.found_location_id and c.found_location_id != a.location_id:
            old = a.location_obj.name if a.location_obj else '—'
            a.location_id = c.found_location_id
            changes.append(f'Local: {old} → {c.found_location.name if c.found_location else "—"}')
        if c.found_condition and c.found_condition != a.asset_condition:
            changes.append(f'Condición: {a.asset_condition} → {c.found_condition}')
            a.asset_condition = c.found_condition
        if changes:
            db.session.add(AssetMovement(
                asset_id=a.id, movement_type='transfer',
                movement_date=(c.counted_at or datetime.utcnow()).date(),
                from_value='Ajuste por inventario físico', to_value=' · '.join(changes),
                notes=f'Conteo: {session.name}',
                created_by=current_user.id))
            applied += 1

    session.applied_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'message': f'{applied} activos actualizados según el conteo',
                    'applied': applied}), 200
