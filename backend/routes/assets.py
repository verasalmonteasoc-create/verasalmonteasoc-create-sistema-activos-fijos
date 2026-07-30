"""
Rutas de Gestión de Activos
"""
from flask import Blueprint, request, jsonify, send_file, send_from_directory
from flask_login import current_user, login_required
from datetime import datetime
from decimal import Decimal
from backend.models import db, Asset, AssetCategory, DepreciationRecord, AuditLog, Department
from io import BytesIO
import os
import openpyxl
from werkzeug.utils import secure_filename
import tempfile

assets_bp = Blueprint('assets', __name__, url_prefix='/api/assets')


def get_asset_qr_url(asset):
    """URL que codifica el QR: la ficha PÚBLICA de solo consulta del activo.

    Apunta a /activo/<id> (no a la aplicación), de modo que escanear un QR
    muestre únicamente la información del activo y nunca dé acceso al sistema.

    Usa APP_BASE_URL si está configurada (obligatorio en producción, ya que
    request.host_url puede resolver a una dirección interna detrás de un
    proxy/balanceador). Si no, cae a la URL de la petición actual.
    """
    base_url = os.environ.get('APP_BASE_URL', '').strip()
    if not base_url:
        base_url = request.host_url
    return f"{base_url.rstrip('/')}/activo/{asset.id}"


def serialize_asset(asset):
    """Convertir Asset a diccionario JSON"""
    department_name = None
    if asset.department_obj:
        department_name = asset.department_obj.name

    location_name = None
    if asset.location_obj:
        location_name = asset.location_obj.name

    return {
        'id': asset.id,
        'code': asset.code,
        'description': asset.description,
        'category': {
            'id': asset.category.id,
            'name': asset.category.name,
            'depreciation_rate': float(asset.category.depreciation_rate)
        },
        'acquisition_date': asset.acquisition_date.isoformat(),
        'acquisition_cost': float(asset.acquisition_cost),
        'residual_value_percent': float(asset.residual_value_percent),
        'useful_life_years': asset.useful_life_years,
        'location': asset.location,
        'location_name': location_name,
        'department': department_name,
        'responsible': asset.responsible,
        'serial_number': asset.serial_number,
        'supplier_name': asset.supplier_name,
        'fiscal_receipt_number': asset.fiscal_receipt_number,
        'acquisition_year': asset.acquisition_year,
        'invoice_filename': asset.invoice_filename,
        'status': asset.status,
        'warranty': asset.warranty,
        'asset_user': asset.asset_user,
        'color': asset.color,
        'year_manufactured': asset.year_manufactured,
        'brand': asset.brand,
        'model': asset.model,
        'chassis': asset.chassis,
        'plate_number': asset.plate_number,
        'equipment_serial': asset.equipment_serial,
        'equipment_supplier': asset.equipment_supplier,
        'physical_location': asset.physical_location,
        'asset_condition': asset.asset_condition,
        'accumulated_depreciation': asset.get_accumulated_depreciation(),
        'net_book_value': asset.get_net_book_value(),
        'last_verified_at': asset.last_verified_at.isoformat() if asset.last_verified_at else None,
        'is_fully_depreciated': asset.is_fully_depreciated(),
        'created_at': asset.created_at.isoformat(),
        'updated_at': asset.updated_at.isoformat()
    }


@assets_bp.route('', methods=['GET'])
def get_assets():
    """Listar activos con filtros"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status')
    category_id = request.args.get('category_id', type=int)
    search = request.args.get('search')

    query = Asset.query

    # Filtros
    if status:
        query = query.filter_by(status=status)
    if category_id:
        query = query.filter_by(category_id=category_id)
    if search:
        query = query.filter(
            (Asset.code.ilike(f'%{search}%')) |
            (Asset.description.ilike(f'%{search}%')) |
            (Asset.serial_number.ilike(f'%{search}%'))
        )

    assets_page = query.paginate(page=page, per_page=per_page)

    return jsonify({
        'success': True,
        'assets': [serialize_asset(a) for a in assets_page.items],
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': assets_page.total,
            'pages': assets_page.pages
        }
    }), 200


@assets_bp.route('/<int:asset_id>', methods=['GET'])
def get_asset(asset_id):
    """Obtener un activo específico"""
    asset = Asset.query.get_or_404(asset_id)
    return jsonify({
        'success': True,
        'asset': serialize_asset(asset)
    }), 200


@assets_bp.route('', methods=['POST'])
def create_asset():
    """Crear nuevo activo"""
    try:
        # Manejo de FormData (con archivos) o JSON
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()

        if not data:
            return jsonify({'success': False, 'message': 'Datos vacíos'}), 400

        invoice_filename = None

        # Manejar archivo de factura si existe
        if 'invoice_file' in request.files:
            file = request.files['invoice_file']
            if file and file.filename:
                try:
                    import os
                    from werkzeug.utils import secure_filename

                    upload_folder = os.path.join(os.path.dirname(__file__), '..', '..', 'uploads', 'invoices')
                    os.makedirs(upload_folder, exist_ok=True)

                    # Guardar con nombre seguro
                    timestamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')
                    filename = secure_filename(f"{timestamp}_{file.filename}")
                    filepath = os.path.join(upload_folder, filename)
                    file.save(filepath)
                    invoice_filename = filename
                except Exception as e:
                    return jsonify({'success': False, 'message': f'Error guardando factura: {str(e)}'}), 400

        # Validación
        required_fields = ['description', 'category_id', 'acquisition_date', 'acquisition_cost', 'useful_life_years']
        if not all(field in data for field in required_fields):
            return jsonify({'success': False, 'message': 'Campos requeridos faltantes'}), 400

        # Verificar categoría
        category = AssetCategory.query.get(data['category_id'])
        if not category:
            return jsonify({'success': False, 'message': 'Categoría no encontrada'}), 404

        # Obtener departamento si existe
        department_id = None
        if data.get('department'):
            dept = Department.query.filter_by(name=data.get('department')).first()
            if dept:
                department_id = dept.id

        # Obtener localidad si existe
        location_id = None
        if data.get('location_name'):
            from backend.models import Location
            loc = Location.query.filter_by(name=data.get('location_name')).first()
            if loc:
                location_id = loc.id

        # Crear activo
        asset = Asset(
            description=data['description'],
            category_id=data['category_id'],
            acquisition_date=datetime.fromisoformat(data['acquisition_date']).date(),
            acquisition_cost=Decimal(str(data['acquisition_cost'])),
            residual_value_percent=Decimal(str(data.get('residual_value_percent', 10))),
            useful_life_years=data['useful_life_years'],
            location=data.get('location', ''),
            location_id=location_id,
            department_id=department_id,
            responsible=data.get('responsible', ''),
            serial_number=data.get('serial_number'),
            supplier_name=data.get('supplier_name', ''),
            fiscal_receipt_number=data.get('fiscal_receipt_number', ''),
            acquisition_year=int(data.get('acquisition_year', datetime.now().year)),
            invoice_filename=invoice_filename,
            status=data.get('status', 'active'),
            warranty=data.get('warranty', ''),
            asset_user=data.get('asset_user', ''),
            color=data.get('color', ''),
            year_manufactured=int(data.get('year_manufactured', 0)) if data.get('year_manufactured') else None,
            brand=data.get('brand', ''),
            model=data.get('model', ''),
            chassis=data.get('chassis', ''),
            plate_number=data.get('plate_number', ''),
            equipment_serial=data.get('equipment_serial', ''),
            equipment_supplier=data.get('equipment_supplier', ''),
            physical_location=data.get('physical_location', ''),
            asset_condition=data.get('asset_condition', 'good'),
            created_by=1,
            notes=data.get('notes', '')
        )

        db.session.add(asset)
        db.session.commit()

        # Registrar en auditoría
        audit = AuditLog(
            user_id=1,
            entity_type='Asset',
            entity_id=asset.id,
            action='create',
            new_value={'description': asset.description, 'code': asset.code},
            description=f'Activo creado: {asset.code} - {asset.description}'
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Activo creado exitosamente',
            'asset': serialize_asset(asset)
        }), 201

    except ValueError as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error de validación: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error creando activo: {str(e)}'}), 500


@assets_bp.route('/<int:asset_id>', methods=['PUT'])
def update_asset(asset_id):
    """Actualizar activo"""
    try:
        asset = Asset.query.get_or_404(asset_id)
        data = request.get_json()

        if not data:
            return jsonify({'success': False, 'message': 'Datos vacíos'}), 400

        # Guardar valores anteriores
        old_values = {}

        # Campos actualizables
        updatable_fields = [
            'description', 'location', 'responsible', 'status', 'notes', 'category_id',
            'acquisition_cost', 'useful_life_years', 'warranty', 'asset_user', 'color',
            'year_manufactured', 'brand', 'model', 'chassis', 'plate_number',
            'equipment_serial', 'equipment_supplier', 'physical_location', 'asset_condition'
        ]
        for field in updatable_fields:
            if field in data:
                old_values[field] = getattr(asset, field)
                if field == 'category_id':
                    # Validar que la categoría exista
                    if not data[field]:
                        return jsonify({'success': False, 'message': 'Categoría requerida'}), 400
                    category = AssetCategory.query.get(data[field])
                    if not category:
                        return jsonify({'success': False, 'message': 'Categoría no encontrada'}), 404
                    setattr(asset, field, data[field])
                elif field == 'acquisition_cost':
                    try:
                        setattr(asset, field, Decimal(str(data[field])))
                    except:
                        return jsonify({'success': False, 'message': 'Costo de adquisición inválido'}), 400
                elif field == 'year_manufactured':
                    try:
                        value = int(data[field]) if data[field] else None
                        setattr(asset, field, value)
                    except:
                        return jsonify({'success': False, 'message': 'Año de manufactura debe ser un número'}), 400
                else:
                    setattr(asset, field, data[field])

        # Manejar departamento especialmente
        if 'department' in data:
            old_values['department'] = asset.department_obj.name if asset.department_obj else None
            if data.get('department'):
                dept = Department.query.filter_by(name=data.get('department')).first()
                if dept:
                    asset.department_id = dept.id
                else:
                    asset.department_id = None
            else:
                asset.department_id = None

        # Manejar localidad especialmente
        if 'location_name' in data:
            old_values['location_name'] = asset.location_obj.name if asset.location_obj else None
            if data.get('location_name'):
                from backend.models import Location
                loc = Location.query.filter_by(name=data.get('location_name')).first()
                if loc:
                    asset.location_id = loc.id
                else:
                    asset.location_id = None
            else:
                asset.location_id = None

        asset.updated_at = datetime.utcnow()
        db.session.commit()

        # Registrar en auditoría
        audit = AuditLog(
            user_id=1,
            entity_type='Asset',
            entity_id=asset.id,
            action='update',
            old_value=old_values,
            new_value={k: data[k] for k in (updatable_fields + ['department', 'location_name']) if k in data},
            description=f'Activo actualizado: {asset.code}'
        )
        db.session.add(audit)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Activo actualizado',
            'asset': serialize_asset(asset)
        }), 200

    except ValueError as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error de validación: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error actualizando activo: {str(e)}'}), 500


@assets_bp.route('/<int:asset_id>', methods=['DELETE'])
def delete_asset(asset_id):
    """Eliminar activo"""
    asset = Asset.query.get_or_404(asset_id)

    # Registrar en auditoría antes de eliminar
    audit = AuditLog(
        user_id=1,
        entity_type='Asset',
        entity_id=asset.id,
        action='delete',
        old_value={'code': asset.code, 'description': asset.description},
        description=f'Activo eliminado: {asset.code}'
    )
    db.session.add(audit)

    db.session.delete(asset)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Activo eliminado'
    }), 200


@assets_bp.route('/<int:asset_id>/depreciation', methods=['GET'])
def get_asset_depreciation(asset_id):
    """Obtener historial de depreciación de un activo"""
    asset = Asset.query.get_or_404(asset_id)

    records = DepreciationRecord.query.filter_by(asset_id=asset_id).order_by(
        DepreciationRecord.year.desc(),
        DepreciationRecord.month.desc()
    ).all()

    return jsonify({
        'success': True,
        'asset': {
            'code': asset.code,
            'description': asset.description,
            'acquisition_cost': float(asset.acquisition_cost),
            'accumulated_depreciation': asset.get_accumulated_depreciation(),
            'net_book_value': asset.get_net_book_value()
        },
        'depreciation_records': [{
            'id': r.id,
            'month_year': r.get_display_month(),
            'depreciation_amount': float(r.depreciation_amount),
            'accumulated_depreciation': float(r.accumulated_depreciation),
            'net_book_value': float(r.net_book_value),
            'calculated_at': r.calculated_at.isoformat()
        } for r in records],
        'total_records': len(records)
    }), 200


@assets_bp.route('/<int:asset_id>/retire', methods=['POST'])
def retire_asset(asset_id):
    """Retiro de activo"""
    asset = Asset.query.get_or_404(asset_id)
    data = request.get_json()

    old_status = asset.status
    asset.status = 'retired'
    asset.notes = data.get('reason', '') + '\nRetirado: ' + datetime.utcnow().isoformat()

    db.session.commit()

    # Auditoría
    audit = AuditLog(
        user_id=1,
        entity_type='Asset',
        entity_id=asset.id,
        action='retire',
        old_value={'status': old_status},
        new_value={'status': 'retired'},
        description=f'Activo retirado: {asset.code}'
    )
    db.session.add(audit)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Activo retirado',
        'asset': serialize_asset(asset)
    }), 200


@assets_bp.route('/<int:asset_id>/summary', methods=['GET'])
def get_asset_summary(asset_id):
    """Obtener resumen del activo"""
    asset = Asset.query.get_or_404(asset_id)

    return jsonify({
        'success': True,
        'summary': {
            'code': asset.code,
            'description': asset.description,
            'category': asset.category.name,
            'acquisition_cost': float(asset.acquisition_cost),
            'residual_value': asset.get_residual_value(),
            'depreciable_amount': asset.get_depreciable_amount(),
            'monthly_depreciation': asset.get_monthly_depreciation(),
            'accumulated_depreciation': asset.get_accumulated_depreciation(),
            'net_book_value': asset.get_net_book_value(),
            'life_remaining_months': max(0, (asset.useful_life_years * 12) - len(
                DepreciationRecord.query.filter_by(asset_id=asset_id).all()
            )),
            'is_fully_depreciated': asset.is_fully_depreciated(),
            'status': asset.status
        }
    }), 200


@assets_bp.route('/<int:asset_id>/invoice', methods=['GET'])
def download_invoice(asset_id):
    """Descargar factura del activo"""
    asset = Asset.query.get_or_404(asset_id)

    if not asset.invoice_filename:
        return jsonify({
            'success': False,
            'message': 'No hay factura adjunta para este activo'
        }), 404

    try:
        upload_folder = os.path.join(os.path.dirname(__file__), '..', '..', 'uploads', 'invoices')
        file_path = os.path.join(upload_folder, asset.invoice_filename)

        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'message': 'Archivo de factura no encontrado'
            }), 404

        return send_file(
            file_path,
            as_attachment=True,
            download_name=f"factura_{asset.code}_{asset.invoice_filename.split('_', 1)[1] if '_' in asset.invoice_filename else asset.invoice_filename}"
        )
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error descargando archivo: {str(e)}'
        }), 500


@assets_bp.route('/<int:asset_id>/invoice/view', methods=['GET'])
def view_invoice(asset_id):
    """Ver factura del activo (no descargar)"""
    asset = Asset.query.get_or_404(asset_id)

    if not asset.invoice_filename:
        return jsonify({
            'success': False,
            'message': 'No hay factura adjunta para este activo'
        }), 404

    try:
        upload_folder = os.path.join(os.path.dirname(__file__), '..', '..', 'uploads', 'invoices')
        file_path = os.path.join(upload_folder, asset.invoice_filename)

        if not os.path.exists(file_path):
            return jsonify({
                'success': False,
                'message': 'Archivo de factura no encontrado'
            }), 404

        return send_file(
            file_path,
            as_attachment=False
        )
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error viewing file: {str(e)}'
        }), 500


@assets_bp.route('/import', methods=['POST'])
def import_assets():
    """Importar activos desde Excel"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No se proporcionó archivo'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Archivo vacío'}), 400

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'El archivo debe ser Excel (.xlsx o .xls)'}), 400

    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file.save(tmp.name)
            temp_path = tmp.name

        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active

        imported_count = 0
        errors = []
        row_num = 2

        # Columnas esperadas: A=Descripción, B=Categoría, C=Costo, D=Fecha, E=Vida Útil, F=Suplidor, G=Factura, H=NCF
        for row in ws.iter_rows(min_row=2, values_only=False):
            try:
                desc = row[0].value if row[0] else None
                cat_name = row[1].value if row[1] else None
                cost = row[2].value if row[2] else None
                acq_date = row[3].value if row[3] else None
                useful_life = row[4].value if row[4] else None
                supplier = row[5].value if row[5] else None
                invoice = row[6].value if row[6] else None
                ncf = row[7].value if row[7] else None

                # Validar campos obligatorios
                if not desc or not cat_name or not cost:
                    errors.append(f'Fila {row_num}: Faltan campos (Descripción, Categoría, Costo)')
                    row_num += 1
                    continue

                # Buscar categoría
                category = AssetCategory.query.filter_by(name=str(cat_name).strip()).first()
                if not category:
                    errors.append(f'Fila {row_num}: Categoría "{cat_name}" no encontrada')
                    row_num += 1
                    continue

                # Crear activo
                asset = Asset(
                    description=str(desc),
                    category_id=category.id,
                    acquisition_cost=Decimal(str(cost)),
                    acquisition_date=acq_date if isinstance(acq_date, datetime) else datetime.now(),
                    useful_life_years=int(useful_life) if useful_life else 5,
                    supplier_name=str(supplier) if supplier else None,
                    fiscal_receipt_number=str(ncf) if ncf else None,
                    status='active'
                )
                db.session.add(asset)
                imported_count += 1

            except Exception as e:
                errors.append(f'Fila {row_num}: {str(e)}')

            row_num += 1

        db.session.commit()
        os.unlink(temp_path)

        return jsonify({
            'success': True,
            'message': f'{imported_count} activos importados',
            'imported_count': imported_count,
            'errors': errors if errors else None
        }), 201

    except Exception as e:
        db.session.rollback()
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.unlink(temp_path)
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 400


@assets_bp.route('/import-auxiliar', methods=['POST'])
def import_auxiliar():
    """Importar el Auxiliar de Activos Fijos (formato con depreciación acumulada).

    Formato esperado (encabezados en cualquiera de las primeras filas):
    MARCA | MODELO | DEPARTAMENTO | COLOR | AÑO | STATUS | CATEGORIA |
    CLASIFICACION | CHASIS | USUARIO | FECHA DE ADQUISICION | COSTO DE
    ADQUISICION | DEPRECIACION ACUMULADA | VALOR EN LIBRO

    REEMPLAZA los activos existentes de las categorías presentes en el archivo
    (el archivo es la fuente de verdad) y registra la depreciación acumulada
    de cada activo como un DepreciationRecord.
    """
    from backend.models import JournalEntry

    if not current_user.is_authenticated or not current_user.is_admin():
        return jsonify({'success': False, 'message': 'Solo un administrador puede importar el auxiliar'}), 403

    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No se proporcionó archivo'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Archivo vacío'}), 400

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'El archivo debe ser Excel (.xlsx o .xls)'}), 400

    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            file.save(tmp.name)
            temp_path = tmp.name

        wb = openpyxl.load_workbook(temp_path, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        # Localizar la fila de encabezados (la que contiene MARCA y MODELO)
        header_idx = None
        for i, row in enumerate(rows[:10]):
            cells = [str(c).strip().upper() if c else '' for c in row[:3]]
            if 'MARCA' in cells:
                header_idx = i
                break
        if header_idx is None:
            return jsonify({
                'success': False,
                'message': 'No se encontró la fila de encabezados (se espera una columna MARCA en las primeras 10 filas)'
            }), 400

        # ── Fase 1: parsear todo el archivo (si algo falla aquí, no se borra nada)
        parsed = []
        errors = []
        for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            try:
                marca = row[0]
                modelo = row[1]
                departamento = row[2]
                color = row[3]
                year = row[4]
                status = row[5]
                categoria = row[6]
                chasis = row[8]
                usuario = row[9]
                fecha_adq = row[10]
                costo = row[11]
                dep_acum = row[12]

                if not costo or not modelo:
                    continue

                if isinstance(fecha_adq, datetime):
                    acq_date = fecha_adq.date()
                elif fecha_adq:
                    acq_date = datetime.strptime(str(fecha_adq).split(' ')[0], '%Y-%m-%d').date()
                else:
                    acq_date = datetime.now().date()

                parsed.append({
                    'marca': str(marca).strip() if marca else '',
                    'modelo': str(modelo).strip() if modelo else '',
                    'departamento': str(departamento).strip() if departamento else None,
                    'color': str(color).strip() if color else '',
                    'year': int(year) if year else None,
                    'status': 'active' if str(status or 'ACTIVO').strip().upper() == 'ACTIVO' else 'inactive',
                    'categoria': str(categoria).strip() if categoria else 'Vehiculos y Camiones Livianos',
                    'chasis': str(chasis).strip() if chasis else '',
                    'usuario': str(usuario).strip() if usuario else '',
                    'acq_date': acq_date,
                    'costo': Decimal(str(costo)),
                    'dep_acum': Decimal(str(dep_acum)) if dep_acum else Decimal('0'),
                })
            except Exception as e:
                errors.append(f'Fila {idx}: {str(e)}')

        if not parsed:
            return jsonify({
                'success': False,
                'message': 'No se encontraron filas válidas para importar',
                'errors': errors[:10]
            }), 400

        # ── Fase 2: asegurar categorías y preparar el upsert (NO se borra nada)
        category_names = sorted({p['categoria'] for p in parsed})
        category_map = {}
        for name in category_names:
            cat = AssetCategory.query.filter_by(name=name).first()
            if not cat:
                cat = AssetCategory(name=name, depreciation_rate=25,
                                    description='Creada por importación de auxiliar')
                db.session.add(cat)
                db.session.flush()
            category_map[name] = cat

        cat_ids = [c.id for c in category_map.values()]

        # Índice de activos existentes por chasis/VIN (clave estable para upsert)
        existing_by_chassis = {}
        for a in Asset.query.filter(Asset.category_id.in_(cat_ids)).all():
            if a.chassis:
                existing_by_chassis[a.chassis.strip().upper()] = a

        # Secuencia para códigos de activos NUEVOS (a partir del máximo VEH-#### existente)
        import re
        max_seq = 0
        for (code,) in db.session.query(Asset.code).all():
            m = re.match(r'VEH-(\d+)$', code or '')
            if m:
                max_seq = max(max_seq, int(m.group(1)))
        next_seq = max_seq + 1

        # ── Fase 3: upsert (actualiza el existente por chasis, crea solo los nuevos)
        created_count = 0
        updated_count = 0
        total_cost = Decimal('0')
        total_dep = Decimal('0')

        for p in parsed:
            dept = None
            if p['departamento']:
                dept = Department.query.filter_by(name=p['departamento']).first()
                if not dept:
                    dept = Department(name=p['departamento'])
                    db.session.add(dept)
                    db.session.flush()

            description = f"{p['marca']} {p['modelo']}".strip()
            chassis_key = p['chasis'].strip().upper() if p['chasis'] else ''
            asset = existing_by_chassis.get(chassis_key) if chassis_key else None

            if asset:
                # Actualizar el activo existente — conserva su ID y sus asientos
                asset.description = description
                asset.category_id = category_map[p['categoria']].id
                asset.acquisition_cost = p['costo']
                asset.acquisition_date = p['acq_date']
                asset.brand = p['marca']
                asset.model = p['modelo']
                asset.color = p['color']
                asset.year_manufactured = p['year']
                asset.asset_user = p['usuario']
                asset.status = p['status']
                if dept:
                    asset.department_id = dept.id
                # El auxiliar es la fuente de verdad del saldo de depreciación:
                # se reemplaza SOLO el saldo de apertura de este activo.
                DepreciationRecord.query.filter_by(asset_id=asset.id).delete(synchronize_session=False)
                updated_count += 1
            else:
                asset = Asset(
                    code=f'VEH-{next_seq:04d}',
                    description=description,
                    category_id=category_map[p['categoria']].id,
                    acquisition_cost=p['costo'],
                    acquisition_date=p['acq_date'],
                    useful_life_years=4,
                    brand=p['marca'],
                    model=p['modelo'],
                    color=p['color'],
                    year_manufactured=p['year'],
                    chassis=p['chasis'],
                    asset_user=p['usuario'],
                    status=p['status'],
                )
                if dept:
                    asset.department_id = dept.id
                db.session.add(asset)
                next_seq += 1
                created_count += 1
                if chassis_key:
                    existing_by_chassis[chassis_key] = asset

            db.session.flush()

            if p['dep_acum'] > 0:
                db.session.add(DepreciationRecord(
                    asset_id=asset.id,
                    year=p['acq_date'].year,
                    month=p['acq_date'].month,
                    depreciation_amount=p['dep_acum'],
                    accumulated_depreciation=p['dep_acum'],
                    net_book_value=p['costo'] - p['dep_acum'],
                    calculated_at=datetime.utcnow(),
                ))

            total_cost += p['costo']
            total_dep += p['dep_acum']

        db.session.commit()

        # Activos de estas categorías que NO vinieron en el archivo (posibles bajas)
        file_chassis = {p['chasis'].strip().upper() for p in parsed if p['chasis']}
        not_in_file = sum(
            1 for a in Asset.query.filter(Asset.category_id.in_(cat_ids)).all()
            if (a.chassis or '').strip().upper() not in file_chassis
        )

        return jsonify({
            'success': True,
            'message': f'{created_count} creados, {updated_count} actualizados',
            'imported_count': created_count + updated_count,
            'created_count': created_count,
            'updated_count': updated_count,
            'not_in_file_count': not_in_file,
            'total_cost': float(total_cost),
            'total_depreciation': float(total_dep),
            'total_net': float(total_cost - total_dep),
            'categories': category_names,
            'errors': errors[:10] if errors else None
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 400
    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


@assets_bp.route('/<int:asset_id>/qrcode', methods=['GET'])
def get_asset_qrcode(asset_id):
    """Generar código QR para un activo"""
    try:
        import qrcode

        asset = Asset.query.get_or_404(asset_id)

        # URL que abre el detalle del activo en modo visualización
        qr_data = get_asset_qr_url(asset)

        # Generar QR
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=8,
            border=2,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Guardar en memoria
        img_io = BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)

        return send_file(
            img_io,
            mimetype='image/png',
            as_attachment=True,
            download_name=f'QR_{asset.code}.png'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error generando QR: {str(e)}'}), 500


@assets_bp.route('/<int:asset_id>/qrcode-data', methods=['GET'])
def get_qrcode_data_uri(asset_id):
    """Retornar QR como data URI (para mostrar en página)"""
    try:
        import qrcode
        import base64

        asset = Asset.query.get_or_404(asset_id)

        # URL que abre el detalle del activo en modo visualización
        qr_data = get_asset_qr_url(asset)

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=8,
            border=2,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Convertir a data URI
        img_io = BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)
        img_base64 = base64.b64encode(img_io.getvalue()).decode('utf-8')
        data_uri = f"data:image/png;base64,{img_base64}"

        return jsonify({
            'success': True,
            'asset_code': asset.code,
            'asset_description': asset.description,
            'qr_data_uri': data_uri
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generando QR: {str(e)}'
        }), 500
