"""
Rutas de Depreciación Mensual
"""
from flask import Blueprint, request, jsonify, current_app
from datetime import datetime
from decimal import Decimal
from backend.models import db, Asset, DepreciationRecord, JournalEntry, JournalEntryLine, AssetCategory, ChartOfAccounts

depreciation_bp = Blueprint('depreciation', __name__, url_prefix='/api/depreciation')


@depreciation_bp.route('/months', methods=['GET'])
def get_processed_months():
    """Obtener lista de meses con depreciación procesada"""
    try:
        # Obtener meses únicos con registros de depreciación
        months_data = db.session.query(
            DepreciationRecord.year,
            DepreciationRecord.month
        ).distinct().order_by(
            DepreciationRecord.year.desc(),
            DepreciationRecord.month.desc()
        ).all()

        months = []
        for year, month in months_data:
            months.append({
                'year': year,
                'month': month,
                'display': f"{_get_month_name(month)} {year}",
                'key': f"{year}-{str(month).zfill(2)}"
            })

        return jsonify({
            'success': True,
            'months': months
        }), 200

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error obteniendo meses: {str(e)}'
        }), 500


@depreciation_bp.route('/process', methods=['POST'])
def process_monthly_depreciation():
    """Procesar depreciación mensual y generar asiento contable"""
    try:
        data = request.get_json()
        year = data.get('year')
        month = data.get('month')
        details = data.get('details', [])

        if not year or not month:
            return jsonify({'success': False, 'message': 'Año y mes requeridos'}), 400

        if not details:
            return jsonify({'success': False, 'message': 'No hay detalles de depreciación'}), 400

        # Verificar si ya existe depreciación para este período
        existing = DepreciationRecord.query.filter(
            DepreciationRecord.year == year,
            DepreciationRecord.month == month
        ).first()

        if existing:
            return jsonify({
                'success': False,
                'message': f'La depreciación para {month}/{year} ya fue procesada. Intenta con un mes diferente.'
            }), 409

        # Agrupar por categoría para generar asientos
        categories_data = {}
        total_depreciation = Decimal('0')

        # Procesar cada activo
        for detail in details:
            asset_id = detail['asset_id']
            monthly_depreciation = Decimal(str(detail['monthlyDepreciation']))
            new_accumulated = Decimal(str(detail['newAccumulated']))

            # Crear registro de depreciación
            net_book_value = Decimal(str(detail['cost'])) - new_accumulated

            dep_record = DepreciationRecord(
                asset_id=asset_id,
                year=year,
                month=month,
                depreciation_amount=monthly_depreciation,
                accumulated_depreciation=new_accumulated,
                net_book_value=net_book_value
            )
            db.session.add(dep_record)

            # Agrupar por categoría
            asset = Asset.query.get(asset_id)
            if asset:
                cat_id = asset.category_id
                if cat_id not in categories_data:
                    categories_data[cat_id] = {
                        'category': asset.category,
                        'depreciation_amount': Decimal('0'),
                        'assets': []
                    }
                categories_data[cat_id]['depreciation_amount'] += monthly_depreciation
                categories_data[cat_id]['assets'].append(asset_id)

            total_depreciation += monthly_depreciation

        db.session.flush()

        # Generar asiento contable único
        journal_entry = JournalEntry(
            reference=f'DEP-{year}{str(month).zfill(2)}',
            description=f'Depreciación de Activos Fijos - {_get_month_name(month)}/{year}',
            entry_date=datetime.now().date(),
            entry_type='depreciation',
            status='posted'
        )
        db.session.add(journal_entry)
        db.session.flush()

        total_debit = Decimal('0')
        total_credit = Decimal('0')

        # Crear líneas de asiento para cada categoría
        for cat_id, cat_data in categories_data.items():
            category = cat_data['category']
            dep_amount = cat_data['depreciation_amount']

            # Obtener cuentas - usar de la categoría o buscar genéricas
            depreciation_expense_acct = category.depreciation_expense_account
            accumulated_depreciation_acct = category.accumulated_depreciation_account

            # Si no están vinculadas, buscar cuentas genéricas
            if not depreciation_expense_acct:
                acct = ChartOfAccounts.query.filter(
                    ChartOfAccounts.account_type == 'Gasto'
                ).first()
                if acct:
                    depreciation_expense_acct = acct.code

            if not accumulated_depreciation_acct:
                acct = ChartOfAccounts.query.filter(
                    ChartOfAccounts.name.ilike('%Deprec%Acumulada%')
                ).first()
                if acct:
                    accumulated_depreciation_acct = acct.code

            # Línea de débito: Gasto de Depreciación
            if depreciation_expense_acct:
                debit_line = JournalEntryLine(
                    journal_entry_id=journal_entry.id,
                    account_code=depreciation_expense_acct,
                    account_name=_get_account_name(depreciation_expense_acct),
                    debit_amount=dep_amount,
                    credit_amount=Decimal('0'),
                    description=f'Depreciación {category.name}'
                )
                db.session.add(debit_line)
                total_debit += dep_amount
            else:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'message': f'No hay cuenta de gasto de depreciación vinculada a {category.name}'
                }), 400

            # Línea de crédito: Depreciación Acumulada
            if accumulated_depreciation_acct:
                credit_line = JournalEntryLine(
                    journal_entry_id=journal_entry.id,
                    account_code=accumulated_depreciation_acct,
                    account_name=_get_account_name(accumulated_depreciation_acct),
                    debit_amount=Decimal('0'),
                    credit_amount=dep_amount,
                    description=f'Depreciación Acumulada {category.name}'
                )
                db.session.add(credit_line)
                total_credit += dep_amount
            else:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'message': f'No hay cuenta de depreciación acumulada vinculada a {category.name}'
                }), 400

        # Actualizar totales del asiento
        journal_entry.total_debit = total_debit
        journal_entry.total_credit = total_credit

        # Validar que debita = créditos
        if total_debit != total_credit:
            db.session.rollback()
            return jsonify({
                'success': False,
                'message': 'Error en balance del asiento contable'
            }), 400

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Depreciación procesada exitosamente',
            'journal_entry_id': journal_entry.id,
            'total_depreciation': float(total_depreciation),
            'assets_processed': len(details),
            'reference': journal_entry.reference
        }), 200

    except ValueError as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Valores inválidos: {str(e)}'
        }), 400
    except KeyError as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Falta campo requerido: {str(e)}'
        }), 400
    except Exception as e:
        db.session.rollback()
        import traceback
        error_trace = traceback.format_exc()
        return jsonify({
            'success': False,
            'message': f'Error procesando depreciación: {str(e)}',
            'error': error_trace if current_app.debug else None
        }), 500


def _get_month_name(month):
    """Obtener nombre del mes en español"""
    months = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    return months.get(month, '')


@depreciation_bp.route('/period-total', methods=['GET'])
def get_period_total():
    """Obtener total de depreciación de un período"""
    try:
        year_from = request.args.get('from', type=str)
        year_to = request.args.get('to', type=str)

        if not year_from or not year_to:
            return jsonify({'success': False, 'message': 'Parámetros requeridos'}), 400

        year_from_int, month_from_int = map(int, year_from.split('-'))
        year_to_int, month_to_int = map(int, year_to.split('-'))

        # Obtener registros del período
        records = db.session.query(DepreciationRecord).filter(
            db.or_(
                db.and_(
                    DepreciationRecord.year == year_from_int,
                    DepreciationRecord.month >= month_from_int
                ),
                db.and_(
                    DepreciationRecord.year == year_to_int,
                    DepreciationRecord.month <= month_to_int
                ),
                db.and_(
                    DepreciationRecord.year > year_from_int,
                    DepreciationRecord.year < year_to_int
                )
            )
        ).all()

        total = sum(Decimal(str(r.depreciation_amount)) for r in records)

        return jsonify({
            'success': True,
            'total_depreciation': float(total)
        }), 200

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@depreciation_bp.route('/period-report', methods=['GET'])
def get_period_report():
    """Generar reporte Excel consolidado de un período"""
    try:
        year_from = request.args.get('year_from', type=int)
        month_from = request.args.get('month_from', type=int)
        year_to = request.args.get('year_to', type=int)
        month_to = request.args.get('month_to', type=int)

        if not all([year_from, month_from, year_to, month_to]):
            return jsonify({'success': False, 'message': 'Parámetros requeridos'}), 400

        # Obtener registros del período
        records = db.session.query(DepreciationRecord).filter(
            db.or_(
                db.and_(
                    DepreciationRecord.year == year_from,
                    DepreciationRecord.month >= month_from
                ),
                db.and_(
                    DepreciationRecord.year == year_to,
                    DepreciationRecord.month <= month_to
                ),
                db.and_(
                    DepreciationRecord.year > year_from,
                    DepreciationRecord.year < year_to
                )
            )
        ).order_by(DepreciationRecord.year, DepreciationRecord.month).all()

        if not records:
            return jsonify({'success': False, 'message': 'No hay registros para este período'}), 404

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "Período"

        # Estilos
        header_fill = PatternFill(start_color="003D7A", end_color="003D7A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws['A1'] = "REPORTE DE DEPRECIACIÓN - PERÍODO CONSOLIDADO"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')

        # Información del período
        start_month = _get_month_name(month_from)
        end_month = _get_month_name(month_to)
        ws['A2'] = f"Período: {start_month} {year_from} a {end_month} {year_to}"
        ws['A3'] = f"Fecha de Reporte: {datetime.now().strftime('%d/%m/%Y')}"

        # Tabla de depreciación
        row = 5
        headers = ["Año-Mes", "Código Activo", "Descripción", "Categoría", "Deprec. Mensual", "Acumulada", "Valor Neto"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        # Datos
        total_depreciation = Decimal('0')
        row = 6
        current_month = None

        for rec in records:
            asset = Asset.query.get(rec.asset_id)
            if asset:
                month_display = f"{rec.year}-{str(rec.month).zfill(2)}"
                ws.cell(row=row, column=1).value = month_display
                ws.cell(row=row, column=2).value = asset.code
                ws.cell(row=row, column=3).value = asset.description
                ws.cell(row=row, column=4).value = asset.category.name if asset.category else ""
                ws.cell(row=row, column=5).value = float(rec.depreciation_amount)
                ws.cell(row=row, column=6).value = float(rec.accumulated_depreciation)
                ws.cell(row=row, column=7).value = float(rec.net_book_value)

                total_depreciation += rec.depreciation_amount
                row += 1

        # Totales
        row += 1
        ws.cell(row=row, column=1).value = "TOTAL DEL PERÍODO"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=5).value = float(total_depreciation)
        ws.cell(row=row, column=5).font = Font(bold=True)

        # Ajustar anchos
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        for col in ['D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18

        # Guardar en memoria
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue(), 200, {
            'Content-Disposition': f'attachment; filename="Depreciacion_{year_from}_{month_from:02d}_a_{year_to}_{month_to:02d}.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@depreciation_bp.route('/report', methods=['GET'])
def get_depreciation_report():
    """Generar reporte Excel de depreciación"""
    try:
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)

        if not year or not month:
            return jsonify({'success': False, 'message': 'Año y mes requeridos'}), 400

        # Obtener registros de depreciación
        dep_records = DepreciationRecord.query.filter_by(year=year, month=month).all()

        if not dep_records:
            return jsonify({'success': False, 'message': 'No hay registros de depreciación para este período'}), 404

        # Obtener asiento contable
        journal = JournalEntry.query.filter_by(year=year, month=month, entry_type='depreciation').first()

        # Crear Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "Depreciación"

        # Estilos
        header_fill = PatternFill(start_color="003D7A", end_color="003D7A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws['A1'] = "REPORTE DE DEPRECIACIÓN MENSUAL"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')

        # Información del período
        ws['A2'] = f"Período: {_get_month_name(month)} {year}"
        ws['A3'] = f"Fecha de Reporte: {datetime.now().strftime('%d/%m/%Y')}"

        # Tabla de depreciación
        row = 5
        headers = ["Código Activo", "Descripción", "Categoría", "Costo", "Deprec. Mensual", "Acumulada Anterior", "Acumulada Nueva", "Valor Neto"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        # Datos
        total_depreciation = Decimal('0')
        row = 6
        for rec in dep_records:
            asset = Asset.query.get(rec.asset_id)
            if asset:
                ws.cell(row=row, column=1).value = asset.code
                ws.cell(row=row, column=2).value = asset.description
                ws.cell(row=row, column=3).value = asset.category.name if asset.category else ""
                ws.cell(row=row, column=4).value = float(asset.acquisition_cost)
                ws.cell(row=row, column=5).value = float(rec.depreciation_amount)
                ws.cell(row=row, column=6).value = float(rec.accumulated_depreciation) - float(rec.depreciation_amount)
                ws.cell(row=row, column=7).value = float(rec.accumulated_depreciation)
                ws.cell(row=row, column=8).value = float(rec.net_book_value)

                total_depreciation += rec.depreciation_amount
                row += 1

        # Totales
        row += 1
        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=5).value = float(total_depreciation)
        ws.cell(row=row, column=5).font = Font(bold=True)

        # Asiento contable
        if journal:
            row += 3
            ws.cell(row=row, column=1).value = "ASIENTO CONTABLE GENERADO"
            ws.cell(row=row, column=1).font = title_font
            ws.merge_cells(f'A{row}:C{row}')

            row += 1
            ws.cell(row=row, column=1).value = f"Referencia: {journal.reference}"
            ws.cell(row=row, column=2).value = f"Estado: {journal.status}"

            row += 2
            headers_journal = ["Cuenta", "Descripción", "Débito", "Crédito"]
            for col, header in enumerate(headers_journal, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            row += 1
            lines = JournalEntryLine.query.filter_by(journal_entry_id=journal.id).all()
            for line in lines:
                ws.cell(row=row, column=1).value = line.account_code
                ws.cell(row=row, column=2).value = line.account_name
                ws.cell(row=row, column=3).value = float(line.debit_amount) if line.debit_amount > 0 else ""
                ws.cell(row=row, column=4).value = float(line.credit_amount) if line.credit_amount > 0 else ""
                row += 1

            row += 1
            ws.cell(row=row, column=1).value = "TOTALES"
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=3).value = float(journal.total_debit)
            ws.cell(row=row, column=3).font = Font(bold=True)
            ws.cell(row=row, column=4).value = float(journal.total_credit)
            ws.cell(row=row, column=4).font = Font(bold=True)

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        for col in ['D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18

        # Guardar en memoria
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue(), 200, {
            'Content-Disposition': f'attachment; filename="Depreciacion_{year}_{str(month).zfill(2)}.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generando reporte: {str(e)}'
        }), 500


def _get_account_name(account_code):
    """Obtener nombre de cuenta desde su código"""
    account = ChartOfAccounts.query.filter_by(code=account_code).first()
    return account.name if account else account_code


@depreciation_bp.route('/asset-entry/<int:asset_id>', methods=['POST'])
def create_asset_journal_entry(asset_id):
    """Crear asiento contable inicial para un activo (compra/adquisición)"""
    try:
        asset = Asset.query.get_or_404(asset_id)

        # Verificar si la categoría tiene cuentas configuradas
        category = asset.category
        if not category.asset_account or not category.accumulated_depreciation_account:
            return jsonify({
                'success': False,
                'message': 'La categoría del activo no tiene cuentas contables configuradas'
            }), 400

        # Buscar las cuentas
        asset_account = ChartOfAccounts.query.filter_by(code=category.asset_account).first()
        if not asset_account:
            return jsonify({
                'success': False,
                'message': f'Cuenta de activo {category.asset_account} no encontrada'
            }), 400

        # Crear asiento contable de compra del activo
        # Débito: Cuenta de Activo
        # Crédito: Caja/Bancos (por ahora dejamos como balancing)
        entry = JournalEntry(
            entry_date=asset.acquisition_date,
            description=f'Compra de Activo: {asset.code} - {asset.description}',
            debit_account_id=asset_account.id,
            amount=asset.acquisition_cost,
            asset_id=asset.id,
            year=asset.acquisition_date.year,
            month=asset.acquisition_date.month,
            entry_type='asset_purchase',
            status='posted',
            total_debit=asset.acquisition_cost,
            total_credit=asset.acquisition_cost
        )

        db.session.add(entry)

        # Crear líneas del asiento
        # Línea de débito - Activo
        line_debit = JournalEntryLine(
            journal_entry=entry,
            account_code=asset_account.code,
            account_name=asset_account.name,
            description=f'{asset.code} - {asset.description}',
            debit_amount=asset.acquisition_cost,
            credit_amount=Decimal('0.00')
        )
        db.session.add(line_debit)

        # Línea de crédito - Contrapartida (asumimos como Cuentas por Pagar)
        line_credit = JournalEntryLine(
            journal_entry=entry,
            account_code='2100',  # Código genérico para cuentas por pagar
            account_name='Cuentas por Pagar',
            description=f'Contrapartida: Compra de {asset.code}',
            debit_amount=Decimal('0.00'),
            credit_amount=asset.acquisition_cost
        )
        db.session.add(line_credit)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Asiento contable creado para el activo {asset.code}',
            'entry': {
                'id': entry.id,
                'entry_date': entry.entry_date.isoformat(),
                'description': entry.description,
                'amount': float(entry.amount),
                'status': entry.status,
                'entry_type': entry.entry_type
            }
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error creando asiento contable: {str(e)}'
        }), 500
